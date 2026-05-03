"""
utils/ea_pollution_inventory.py — Environment Agency Pollution Inventory.

The EA's Pollution Inventory (https://data.gov.uk/dataset/pollution-inventory)
is the UK PRTR — annual point-source emissions for every industrial site
permitted under the Environmental Permitting Regulations. One workbook per
year (2013–2024) with three sheets:

    "<year> Substances"        — emissions to air / water / land / sewer
    "<year> Waste Transfers"   — wastes transferred off-site
    "<year> Radioactive Wastes"

Per-row columns (Substances sheet):
    Permit Number | Activity | Operator | Address | Postcode |
    Easting | Northing | Region | Medium | Substance | Threshold (kg) |
    Quantity (kg) | Activity Category | Site Type

Public surface:
    ensure_year(year) -> Path        download + extract once, returns XLSX
    load_substances(year) -> list    cached parsed rows (~31 k for 2024)
    by_operator(query, year) -> list
    by_postcode(postcode, year) -> list
    nearest(lat, lon, radius_km, year)
    asset_emissions_summary(name|operator, lat|lon, year)

Designed to be cheap on the first call (~5 MB download), fast after — the
parsed dataset is held in a module-level dict keyed by year.
"""

from __future__ import annotations

import logging
import math
import os
import zipfile
from dataclasses import dataclass, asdict
from functools import lru_cache
from pathlib import Path
from typing import Any
from urllib.parse import quote

import httpx

log = logging.getLogger("princeps.ea_pollution_inventory")

_CACHE_DIR = Path(__file__).resolve().parent.parent / "data" / "ea_epr"
_CACHE_DIR.mkdir(parents=True, exist_ok=True)

_USER_AGENT = "Princeps/1.0 (+https://princeps.energy) ea-pi"
_TIMEOUT = httpx.Timeout(120.0, connect=15.0)

_FILE_DATASET_ID = "4faa4a52-7df2-4047-bc3f-877dd04222d8"
# Year → public file name (the EA renamed each release inconsistently, so we
# keep an explicit map rather than guess).
_FILE_NAME_BY_YEAR: dict[int, str] = {
    2013: "2013_Pollution_Inventory.zip",
    2014: "2014_Pollution_Inventory.zip",
    2015: "2015_Pollution_Inventory.zip",
    2016: "2016_Pollution_Inventory_dataset_-_version_2.zip",
    2017: "2017_Pollution_Inventory_Dataset.zip",
    2018: "2018_Pollution_Inventory_Version_1.zip",
    2019: "2019 Pollution Inventory Dataset v2.zip",
    2020: "2020_Pollution_Inventory_Dataset_v2.zip",
    2021: "2021 Pollution Inventory Dataset.zip",
    2022: "2022 Pollution Inventory Dataset v2.zip",
    2023: "2023 Pollution Inventory Dataset V2.zip",
    2024: "2024 Pollution Inventory Dataset.zip",
}

LATEST_YEAR = max(_FILE_NAME_BY_YEAR.keys())


@dataclass(frozen=True)
class PiRow:
    permit_number: str
    activity: str | None
    operator: str | None
    address: str | None
    postcode: str | None
    easting: float | None
    northing: float | None
    region: str | None
    medium: str | None
    substance: str | None
    threshold_kg: float | None
    quantity_kg: float | None
    activity_category: str | None
    site_type: str | None
    year: int


# ---------------------------------------------------------------------------
# Download + extract
# ---------------------------------------------------------------------------
async def ensure_year(year: int = LATEST_YEAR, *, force: bool = False) -> Path:
    """Download + extract the year's XLSX (cached on disk). Returns the path."""
    fname = _FILE_NAME_BY_YEAR.get(year)
    if fname is None:
        raise ValueError(f"unsupported PI year {year}; valid: {sorted(_FILE_NAME_BY_YEAR)}")
    zip_path = _CACHE_DIR / f"pi{year}.zip"
    xlsx_path = _CACHE_DIR / f"pi{year}.xlsx"
    if xlsx_path.exists() and not force:
        return xlsx_path

    if not zip_path.exists() or force:
        async with httpx.AsyncClient(
            timeout=_TIMEOUT, headers={"User-Agent": _USER_AGENT},
            follow_redirects=True,
        ) as c:
            r = await c.get(
                "https://environment.data.gov.uk/api/file/download",
                params={"fileDataSetId": _FILE_DATASET_ID, "fileName": fname},
            )
            r.raise_for_status()
            zip_path.write_bytes(r.content)
            log.info("EA PI %d downloaded: %d bytes", year, len(r.content))

    with zipfile.ZipFile(zip_path) as z:
        for name in z.namelist():
            if name.lower().endswith(".xlsx"):
                with z.open(name) as src, open(xlsx_path, "wb") as dst:
                    dst.write(src.read())
                break
    return xlsx_path


# ---------------------------------------------------------------------------
# Parse — substances sheet only (waste / radioactive available on demand)
# ---------------------------------------------------------------------------
_SUBSTANCES_CACHE: dict[int, list[PiRow]] = {}


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_substances(year: int = LATEST_YEAR) -> list[PiRow]:
    if year in _SUBSTANCES_CACHE:
        return _SUBSTANCES_CACHE[year]
    xlsx_path = _CACHE_DIR / f"pi{year}.xlsx"
    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"PI {year} workbook missing — call ensure_year({year}) first"
        )
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sn = next(
        (s for s in wb.sheetnames if "substance" in s.lower() and str(year) in s),
        None,
    )
    if sn is None:
        log.warning("no substances sheet in PI %d", year)
        _SUBSTANCES_CACHE[year] = []
        return []
    ws = wb[sn]
    rows: list[PiRow] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Real data starts after a 4-line header block; first cell is the
        # permit number, which is alphanumeric without spaces (e.g. BV3006IN).
        if not row or not row[0]:
            continue
        first = str(row[0]).strip()
        if i < 4 or " " in first or len(first) > 16:
            continue
        rows.append(PiRow(
            permit_number=first,
            activity=str(row[1]).strip() if row[1] else None,
            operator=str(row[2]).strip() if row[2] else None,
            address=str(row[3]).strip() if row[3] else None,
            postcode=str(row[4]).strip() if row[4] else None,
            easting=_to_float(row[5]),
            northing=_to_float(row[6]),
            region=str(row[7]).strip() if row[7] else None,
            medium=str(row[8]).strip() if row[8] else None,
            substance=str(row[9]).strip() if row[9] else None,
            threshold_kg=_to_float(row[10]),
            quantity_kg=_to_float(row[11]),
            activity_category=str(row[12]).strip() if row[12] else None,
            site_type=str(row[13]).strip() if len(row) > 13 and row[13] else None,
            year=year,
        ))
    _SUBSTANCES_CACHE[year] = rows
    log.info("EA PI %d loaded: %d substance rows", year, len(rows))
    return rows


# ---------------------------------------------------------------------------
# Lookups
# ---------------------------------------------------------------------------
def _norm(s: str | None) -> str:
    if not s:
        return ""
    return "".join(ch.lower() for ch in s if ch.isalnum())


def _osgb_to_wgs84(e: float, n: float) -> tuple[float, float] | None:
    """Approximate OSGB36 (m) → WGS84 (lat, lon). Uses pyproj when available;
    otherwise returns None — caller falls back to Haversine using OSGB metric."""
    try:
        from pyproj import Transformer
        t = Transformer.from_crs("EPSG:27700", "EPSG:4326", always_xy=True)
        lon, lat = t.transform(e, n)
        return (lat, lon)
    except Exception:
        return None


def by_operator(query: str, year: int = LATEST_YEAR) -> list[PiRow]:
    """Operator OR address substring match (case-insensitive, normalised)."""
    if not query:
        return []
    q = _norm(query)
    if len(q) < 3:
        return []
    out = []
    for r in load_substances(year):
        if (q in _norm(r.operator)
            or q in _norm(r.address)
            or q in _norm(r.permit_number)):
            out.append(r)
    return out


def by_postcode(postcode: str, year: int = LATEST_YEAR) -> list[PiRow]:
    p = _norm(postcode)
    if not p:
        return []
    return [r for r in load_substances(year) if _norm(r.postcode) == p]


def nearest(
    lat: float,
    lon: float,
    radius_km: float = 5.0,
    year: int = LATEST_YEAR,
    max_n: int = 10,
) -> list[tuple[float, PiRow]]:
    """Spatial nearest by easting/northing → 27700 metric distance."""
    try:
        from pyproj import Transformer
        t = Transformer.from_crs("EPSG:4326", "EPSG:27700", always_xy=True)
        e0, n0 = t.transform(lon, lat)
    except Exception:
        log.warning("pyproj unavailable — nearest() spatial filter disabled")
        return []
    found: dict[str, tuple[float, PiRow]] = {}
    for r in load_substances(year):
        if r.easting is None or r.northing is None:
            continue
        d_m = math.hypot(r.easting - e0, r.northing - n0)
        if d_m > radius_km * 1000:
            continue
        d_km = d_m / 1000.0
        key = r.permit_number
        if key not in found or d_km < found[key][0]:
            found[key] = (d_km, r)
    return sorted(found.values(), key=lambda kv: kv[0])[:max_n]


# ---------------------------------------------------------------------------
# Convenience for asset-intel popup
# ---------------------------------------------------------------------------
async def asset_emissions_summary(
    *,
    name: str | None = None,
    operator: str | None = None,
    lat: float | None = None,
    lon: float | None = None,
    radius_km: float = 1.0,
    year: int = LATEST_YEAR,
) -> dict | None:
    """Resolve to one site (best-match permit) and bucket emissions per medium."""
    await ensure_year(year)

    candidates: list[PiRow] = []
    site_label = name or operator
    if name:
        candidates = by_operator(name, year=year)
    if not candidates and operator:
        candidates = by_operator(operator, year=year)
    if not candidates and lat is not None and lon is not None:
        spatial = nearest(lat, lon, radius_km=radius_km, year=year, max_n=8)
        # nearest already deduped — flatten to rows
        for d_km, r in spatial:
            extras = [x for x in load_substances(year) if x.permit_number == r.permit_number]
            candidates.extend(extras)
            break    # keep the closest permit only

    if not candidates:
        return None

    # Group by permit_number; one site may appear under multiple operators
    # if the permit transferred — pick the permit with most rows.
    by_permit: dict[str, list[PiRow]] = {}
    for r in candidates:
        by_permit.setdefault(r.permit_number, []).append(r)
    permit, rows = max(by_permit.items(), key=lambda kv: len(kv[1]))

    # Aggregate by medium → substance
    by_medium: dict[str, list[dict]] = {}
    for r in rows:
        bucket = by_medium.setdefault(r.medium or "Unknown", [])
        bucket.append({
            "substance": r.substance,
            "quantity_kg": r.quantity_kg,
            "threshold_kg": r.threshold_kg,
            "activity_category": r.activity_category,
        })

    head = rows[0]
    coords = (
        _osgb_to_wgs84(head.easting, head.northing)
        if head.easting is not None and head.northing is not None
        else None
    )

    return {
        "year": year,
        "permit_number": permit,
        "operator": head.operator,
        "address": head.address,
        "postcode": head.postcode,
        "region": head.region,
        "site_type": head.site_type,
        "activity": head.activity,
        "easting": head.easting,
        "northing": head.northing,
        "coordinates_latlng": list(coords) if coords else None,
        "n_rows": len(rows),
        "by_medium": {
            m: sorted(items, key=lambda d: -(d["quantity_kg"] or 0))[:8]
            for m, items in by_medium.items()
        },
        "totals_kg": {
            m: round(sum((d["quantity_kg"] or 0) for d in items), 2)
            for m, items in by_medium.items()
        },
        "source": "Environment Agency Pollution Inventory",
    }
