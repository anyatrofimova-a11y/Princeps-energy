"""
XLSX Financial Export — generates a multi-tab Excel workbook
with site assessment data pre-populated for developer financial models.

Tabs:
  1. Summary — verdict, key metrics, site description
  2. Yield — monthly/annual generation, capacity factor, losses
  3. Grid — nearest substation, headroom, connection cost, power flow
  4. Financials — CAPEX/OPEX benchmarks, revenue streams, NPV/IRR inputs
  5. Constraints — planning, environmental, regulatory flags
  6. Satellite — GeeFlow land use, terrain, solar resource, NDVI
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# ── Styles ──────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C2C3A", end_color="2C2C3A", fill_type="solid")
LABEL_FONT = Font(name="Calibri", bold=True, size=10, color="374151")
VALUE_FONT = Font(name="Calibri", size=10, color="1F2937")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="7C5CFC")
SECTION_FILL = PatternFill(start_color="F3F0FF", end_color="F3F0FF", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="E5E7EB"),
)
GO_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
CAUTION_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
NOGO_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")


def _add_header_row(ws, row, labels):
    for col, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _add_kv(ws, row, label, value, col=1):
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = LABEL_FONT
    lc.border = THIN_BORDER
    vc = ws.cell(row=row, column=col + 1, value=value)
    vc.font = VALUE_FONT
    vc.border = THIN_BORDER
    return row + 1


def _add_section(ws, row, title, col=1, span=2):
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(col, col + span):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    return row + 1


def generate_xlsx(data: dict) -> bytes:
    """
    Generate an XLSX workbook from site assessment data.

    Args:
        data: dict with keys: site, verdict, yield_data, grid, financials,
              constraints, satellite, placed_assets
    Returns:
        bytes of the XLSX file
    """
    wb = openpyxl.Workbook()

    # ═══ Tab 1: Summary ═══
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    r = 1
    ws.cell(row=r, column=1, value="PRINCEPS — Site Assessment").font = Font(name="Calibri", bold=True, size=14, color="7C5CFC")
    r = 2
    ws.cell(row=r, column=1, value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}").font = Font(size=9, color="9CA3AF")
    r = 4

    site = data.get("site", {})
    r = _add_section(ws, r, "SITE")
    r = _add_kv(ws, r, "Name", site.get("name", "—"))
    r = _add_kv(ws, r, "Latitude", site.get("lat"))
    r = _add_kv(ws, r, "Longitude", site.get("lon"))
    r = _add_kv(ws, r, "Parcel ID", site.get("parcel_id", "—"))
    r = _add_kv(ws, r, "Area (ha)", site.get("area_ha"))
    r = _add_kv(ws, r, "Capacity (kW)", site.get("capacity_kw"))
    r += 1

    verdict = data.get("verdict", {})
    r = _add_section(ws, r, "VERDICT")
    vc = ws.cell(row=r, column=2, value=verdict.get("verdict", "—"))
    vc.font = Font(bold=True, size=14)
    vfill = GO_FILL if verdict.get("verdict") == "GO" else CAUTION_FILL if verdict.get("verdict") == "CAUTION" else NOGO_FILL
    vc.fill = vfill
    ws.cell(row=r, column=1, value="Verdict").font = LABEL_FONT
    r += 1
    r = _add_kv(ws, r, "Confidence", f"{round((verdict.get('confidence', 0)) * 100)}%")
    r = _add_kv(ws, r, "Summary", verdict.get("summary", "—"))
    r += 1

    # Risks
    risks = verdict.get("risks", [])
    if risks:
        r = _add_section(ws, r, "KEY RISKS")
        for risk in risks[:5]:
            ws.cell(row=r, column=1, value="⚠").font = Font(color="DC2626")
            ws.cell(row=r, column=2, value=risk).font = VALUE_FONT
            r += 1
        r += 1

    # ═══ Tab 2: Yield ═══
    ws2 = wb.create_sheet("Yield")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 18

    yld = data.get("yield_data", {})
    r = 1
    r = _add_section(ws2, r, "ANNUAL YIELD")
    r = _add_kv(ws2, r, "Annual Energy (kWh)", yld.get("annual_energy_kwh"))
    r = _add_kv(ws2, r, "Annual Energy (MWh)", round(yld.get("annual_energy_kwh", 0) / 1000, 1) if yld.get("annual_energy_kwh") else None)
    r = _add_kv(ws2, r, "Capacity Factor (%)", yld.get("capacity_factor_pct"))
    r = _add_kv(ws2, r, "Peak Sun Hours", yld.get("peak_sun_hours"))
    r = _add_kv(ws2, r, "System Losses (%)", yld.get("system_losses_pct"))
    r += 1

    monthly = yld.get("monthly_energy_kwh", [])
    if monthly:
        r = _add_section(ws2, r, "MONTHLY GENERATION")
        _add_header_row(ws2, r, ["Month", "Energy (kWh)", "Energy (MWh)"])
        r += 1
        months = ["January", "February", "March", "April", "May", "June",
                   "July", "August", "September", "October", "November", "December"]
        for i, kwh in enumerate(monthly):
            ws2.cell(row=r, column=1, value=months[i] if i < 12 else f"M{i+1}").font = VALUE_FONT
            ws2.cell(row=r, column=2, value=kwh).font = VALUE_FONT
            ws2.cell(row=r, column=3, value=round(kwh / 1000, 1)).font = VALUE_FONT
            r += 1

    # ═══ Tab 3: Grid ═══
    ws3 = wb.create_sheet("Grid")
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 30

    grid = data.get("grid", {})
    r = 1
    r = _add_section(ws3, r, "NEAREST CONNECTION POINT")
    r = _add_kv(ws3, r, "Substation", grid.get("nearest_substation"))
    r = _add_kv(ws3, r, "Distance (km)", grid.get("distance_km"))
    r = _add_kv(ws3, r, "Voltage (kV)", grid.get("voltage_kv"))
    r = _add_kv(ws3, r, "Headroom (MW)", grid.get("headroom_mw"))
    r = _add_kv(ws3, r, "DNO", grid.get("dno"))
    r += 1

    r = _add_section(ws3, r, "CONNECTION COST ESTIMATE")
    costs = grid.get("connection_cost", {})
    r = _add_kv(ws3, r, "P10 (optimistic)", f"£{costs.get('p10_gbp', 0):,.0f}" if costs.get("p10_gbp") else "—")
    r = _add_kv(ws3, r, "P50 (expected)", f"£{costs.get('p50_gbp', 0):,.0f}" if costs.get("p50_gbp") else "—")
    r = _add_kv(ws3, r, "P90 (conservative)", f"£{costs.get('p90_gbp', 0):,.0f}" if costs.get("p90_gbp") else "—")
    r += 1

    r = _add_section(ws3, r, "UK CONNECTION RATES (£/km)")
    r = _add_kv(ws3, r, "11kV", "£80,000/km")
    r = _add_kv(ws3, r, "33kV", "£150,000/km")
    r = _add_kv(ws3, r, "132kV", "£500,000/km")

    # ═══ Tab 4: Financials ═══
    ws4 = wb.create_sheet("Financials")
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 22
    ws4.column_dimensions["C"].width = 22

    fin = data.get("financials", {})
    r = 1
    r = _add_section(ws4, r, "CAPEX BENCHMARKS (2024-25)")
    _add_header_row(ws4, r, ["Technology", "Low (£/kW)", "High (£/kW)"])
    r += 1
    capex = [
        ("Solar PV (ground-mount)", 450, 950),
        ("Battery Storage (BESS)", 300, 700),
        ("Wind (onshore)", 900, 1800),
        ("EV Charging", 800, 2000),
        ("Transformer", 50, 150),
    ]
    for tech, low, high in capex:
        ws4.cell(row=r, column=1, value=tech).font = VALUE_FONT
        ws4.cell(row=r, column=2, value=low).font = VALUE_FONT
        ws4.cell(row=r, column=3, value=high).font = VALUE_FONT
        r += 1
    r += 1

    r = _add_section(ws4, r, "REVENUE ASSUMPTIONS")
    r = _add_kv(ws4, r, "Export Rate (p/kWh)", fin.get("export_rate_p", 5.5))
    r = _add_kv(ws4, r, "Import Rate (p/kWh)", fin.get("import_rate_p", 28.6))
    r = _add_kv(ws4, r, "PPA Rate (£/MWh)", fin.get("ppa_rate_gbp_mwh", 55))
    r = _add_kv(ws4, r, "WACC (%)", fin.get("wacc_pct", 6.0))
    r = _add_kv(ws4, r, "Project Life (years)", fin.get("project_life", 25))
    r = _add_kv(ws4, r, "Degradation (%/yr)", fin.get("degradation_pct", 0.5))
    r += 1

    r = _add_section(ws4, r, "BESS REVENUE STACKING (£/MW/yr)")
    bess_rev = [
        ("FFR Dynamic", "55-75K"),
        ("Dynamic Containment", "35-48K"),
        ("Capacity Market", "12.5K"),
        ("Wholesale Arbitrage", "18-25K"),
        ("DNO Peak Shaving", "22K"),
    ]
    for stream, value in bess_rev:
        r = _add_kv(ws4, r, stream, f"£{value}")

    # ═══ Tab 5: Constraints ═══
    ws5 = wb.create_sheet("Constraints")
    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 40

    cons = data.get("constraints", {})
    r = 1
    r = _add_section(ws5, r, "PLANNING & ENVIRONMENTAL")

    domains = cons.get("domains", {})
    for domain_key, domain_data in domains.items():
        label = domain_key.replace("_", " ").title()
        status = domain_data.get("status", "—")
        detail = domain_data.get("detail", "")
        score = domain_data.get("score", 0)
        ws5.cell(row=r, column=1, value=label).font = LABEL_FONT
        sc = ws5.cell(row=r, column=2, value=f"{status} ({score}/100) — {detail}")
        sc.font = VALUE_FONT
        if status == "GO":
            sc.fill = GO_FILL
        elif status == "CAUTION":
            sc.fill = CAUTION_FILL
        elif status == "NO-GO":
            sc.fill = NOGO_FILL
        r += 1

    r += 1
    r = _add_section(ws5, r, "REGULATORY FRAMEWORKS")
    frameworks = ["G99/G100", "CDM 2015", "BNG (10% uplift)", "EIA Screening", "NPPF", "ALC", "Flood Test", "NSIP (solar >100MW; onshore wind >100MW)"]
    for fw in frameworks:
        r = _add_kv(ws5, r, fw, "Check required")

    # ═══ Tab 6: Satellite ═══
    ws6 = wb.create_sheet("Satellite")
    ws6.column_dimensions["A"].width = 24
    ws6.column_dimensions["B"].width = 30

    sat = data.get("satellite", {})
    r = 1
    r = _add_section(ws6, r, "TERRAIN")
    terrain = sat.get("terrain", {})
    r = _add_kv(ws6, r, "Mean Elevation (m)", terrain.get("elevation_mean_m"))
    r = _add_kv(ws6, r, "Mean Slope (°)", terrain.get("slope_mean_deg"))
    r = _add_kv(ws6, r, "South-facing (%)", terrain.get("south_facing_pct"))
    r = _add_kv(ws6, r, "Roughness", terrain.get("roughness_mean"))
    r += 1

    r = _add_section(ws6, r, "LAND USE (DynamicWorld)")
    land_use = sat.get("land_use", {})
    r = _add_kv(ws6, r, "Dominant Class", land_use.get("dominant_class"))
    r = _add_kv(ws6, r, "Developable (%)", land_use.get("developable_pct"))
    for cls, pct in (land_use.get("class_percentages") or {}).items():
        if pct > 1:
            r = _add_kv(ws6, r, f"  {cls}", f"{pct}%")
    r += 1

    r = _add_section(ws6, r, "SOLAR RESOURCE (ERA5)")
    solar = sat.get("solar_resource", {})
    r = _add_kv(ws6, r, "Annual GHI (kWh/m²)", solar.get("annual_ghi_kwh_m2"))
    r += 1

    r = _add_section(ws6, r, "FLOOD RISK")
    flood = sat.get("flood_risk", {})
    r = _add_kv(ws6, r, "Risk Level", flood.get("risk_level"))
    r = _add_kv(ws6, r, "Water Occurrence (%)", flood.get("water_occurrence_pct"))

    # ═══ Tab 7: Placed Assets ═══
    assets = data.get("placed_assets", [])
    if assets:
        ws7 = wb.create_sheet("Assets")
        ws7.column_dimensions["A"].width = 20
        ws7.column_dimensions["B"].width = 16
        ws7.column_dimensions["C"].width = 12
        ws7.column_dimensions["D"].width = 14
        ws7.column_dimensions["E"].width = 14

        _add_header_row(ws7, 1, ["Asset Type", "Label", "MW", "Latitude", "Longitude"])
        for i, asset in enumerate(assets, 2):
            ws7.cell(row=i, column=1, value=asset.get("assetType", "")).font = VALUE_FONT
            ws7.cell(row=i, column=2, value=asset.get("label", "")).font = VALUE_FONT
            ws7.cell(row=i, column=3, value=asset.get("mw")).font = VALUE_FONT
            ws7.cell(row=i, column=4, value=asset.get("lat")).font = VALUE_FONT
            ws7.cell(row=i, column=5, value=asset.get("lon")).font = VALUE_FONT

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
