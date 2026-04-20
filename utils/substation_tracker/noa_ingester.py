"""NESO NOA + CSNP ingester.

NESO publishes two related documents that bear directly on transmission-level
substation build:

* **NOA** — Network Options Assessment. Annual (usually Jan-Feb release).
  Published as a PDF report + spreadsheet workbook listing every transmission
  reinforcement option, its recommendation (Proceed / Hold / Do Not Start),
  cost estimate (£m, NOA price base), and expected Energy Not Supplied.
* **CSNP** — Centralised Strategic Network Plan (replaces NOA from 2025+
  under REMA). Same shape, different publication name and expanded scope to
  cover offshore + interconnector integration.

Both are sourced from the NESO data portal:
    https://www.neso.energy/data-portal

V1 status
---------
SCAFFOLD — HTTP client + workbook parser in place, but the exact column
mapping is stubbed with the NOA 2024 column layout as the reference. Needs
tuning against NOA/CSNP 2026 once ship lands (expected May 2026 per the 2026
regulatory sweep).

Data portal shape
-----------------
NESO data portal is CKAN — the programmatic API is:

    GET https://www.neso.energy/data/api/3/action/package_show?id=<id>
    GET https://www.neso.energy/data/api/3/action/resource_show?id=<id>

and resource downloads under:

    https://www.neso.energy/data/dataset/<slug>/resource/<uuid>/download/<file>

Known package IDs (stale — verify at ingest time):
    * noa-2024
    * noa-refresh-2024
    * csnp-2025-pathway-to-2030
"""

from __future__ import annotations

import io
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from .schema import (
    SubstationProject,
    PARENT_COMPANY_BY_DEVELOPER,
    COUNTRY_BY_DEVELOPER,
)

log = logging.getLogger("princeps.substation_tracker.noa")

NESO_CKAN_BASE = "https://www.neso.energy/data/api/3/action"

# Known NOA/CSNP CKAN package IDs — update when NESO changes slugs.
NESO_PACKAGES: dict[str, str] = {
    "NOA_2024":   "noa-2024",
    "NOA_REFRESH_2024": "noa-refresh-2024",
    "CSNP_2025":  "csnp-2025-pathway-to-2030",
    # Follow-up: CSNP 2026 expected (reg sweep log).
}


# ---------------------------------------------------------------------------
# Fetch layer
# ---------------------------------------------------------------------------
async def fetch_neso_package(package_id: str) -> dict | None:
    """Call CKAN `package_show` to list a dataset's resources."""
    import httpx
    url = f"{NESO_CKAN_BASE}/package_show?id={package_id}"
    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.get(url)
        if r.status_code != 200:
            log.warning("NESO CKAN %s HTTP %d", package_id, r.status_code)
            return None
        payload = r.json()
        if not payload.get("success"):
            log.warning("NESO CKAN %s unsuccessful response: %s", package_id, payload.get("error"))
            return None
        return payload.get("result")


async def download_resource(url: str, dest: Path) -> Path | None:
    """Download one CKAN resource to a local file."""
    import httpx
    dest.parent.mkdir(parents=True, exist_ok=True)
    async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
        r = await client.get(url)
        if r.status_code != 200:
            log.warning("download %s HTTP %d", url, r.status_code)
            return None
        dest.write_bytes(r.content)
        return dest


# ---------------------------------------------------------------------------
# Top-level orchestrator
# ---------------------------------------------------------------------------
async def ingest_noa(
    download_dir: Path = Path("/Users/anyatrofimova/Downloads/noa"),
    packages: tuple[str, ...] = ("NOA_REFRESH_2024", "CSNP_2025"),
) -> list[SubstationProject]:
    """Pull the latest NOA + CSNP files from NESO and extract substation rows.

    This is a best-effort fetch — it logs and continues on failure so that
    pipelines relying on a partial NOA ingest still get LTDS data.

    Returns the combined SubstationProject list (may be empty).
    """
    results: list[SubstationProject] = []
    for pkg_key in packages:
        pkg_id = NESO_PACKAGES.get(pkg_key, pkg_key)
        meta = await fetch_neso_package(pkg_id)
        if not meta:
            continue
        # Filter resources to Excel workbooks (the project list usually lives there).
        resources = [r for r in (meta.get("resources") or []) if _is_workbook(r)]
        for res in resources:
            url = res.get("url") or res.get("download_url")
            if not url:
                continue
            fname = re.sub(r"[^a-zA-Z0-9._-]+", "_", res.get("name") or Path(url).name)
            dest = download_dir / f"{pkg_key}__{fname}"
            got = await download_resource(url, dest)
            if not got:
                continue
            source_type = "NOA" if pkg_key.startswith("NOA") else "CSNP"
            rows = parse_noa_workbook(got, source_type=source_type, package_id=pkg_id)
            results.extend(rows)
    log.info("NESO NOA ingest: produced %d SubstationProject records", len(results))
    return results


def _is_workbook(resource: dict) -> bool:
    fmt = (resource.get("format") or "").lower()
    name = (resource.get("name") or "").lower()
    return fmt in ("xlsx", "xls", "csv") or name.endswith((".xlsx", ".xls", ".csv"))


# ---------------------------------------------------------------------------
# Parse layer — Excel/CSV workbook → SubstationProject
# ---------------------------------------------------------------------------
# Known NOA 2024 column aliases — update when NOA/CSNP 2026 lands.
NOA_COLUMN_ALIASES = {
    "scheme_name":     ("project name", "option name", "reinforcement name"),
    "description":     ("description", "option description"),
    "to":              ("to", "transmission owner", "licensee"),
    "recommendation":  ("recommendation", "neso recommendation", "decision"),
    "capex_gbp_m":     ("cost estimate", "total cost (£m)", "capex (£m)", "cost (£m)"),
    "earliest_ready":  ("earliest in-service", "earliest ready", "ead", "completion"),
    "voltage_kv":      ("voltage kv", "voltage"),
    "region":          ("region", "zone", "boundary"),
    "noa_id":          ("option id", "noa id", "project id", "reference"),
}

# NOA recommendation → construction_status mapping.
NOA_REC_STATUS = {
    "proceed":             "planned",
    "proceed with plans":  "planned",
    "hold":                "deferred",
    "do not start":        "cancelled",
    "stop":                "cancelled",
    "complete":            "complete",
    "optimised":           "planned",
}


def parse_noa_workbook(
    path: Path,
    source_type: str = "NOA",
    package_id: str | None = None,
) -> list[SubstationProject]:
    """Parse a NOA/CSNP workbook into SubstationProject records.

    Returns only rows that look like substations (not pure OHL or
    interconnector projects). We detect substations via keyword triggers
    in the project name/description.
    """
    if not path.exists():
        log.warning("NOA workbook missing: %s", path)
        return []

    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _parse_csv(path)
    else:
        try:
            import openpyxl  # type: ignore
        except ImportError:
            log.error("openpyxl needed for NOA xlsx")
            return []
        rows = _parse_xlsx(path)

    out: list[SubstationProject] = []
    for rec in rows:
        name = (rec.get("scheme_name") or "").strip() if rec.get("scheme_name") else ""
        descr = (rec.get("description") or "").strip() if rec.get("description") else ""
        if not name or not _has_substation_keyword(name, descr):
            continue
        proj = _noa_row_to_project(rec, source_type=source_type, package_id=package_id, source_file=path.name)
        if proj:
            out.append(proj)
    log.info("NOA parser %s: produced %d substation projects from %d rows", path.name, len(out), len(rows))
    return out


def _parse_xlsx(path: Path) -> list[dict]:
    import openpyxl  # type: ignore
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_rows: list[dict] = []
    for sheet in wb.worksheets:
        header_row, header_map = _find_noa_header(sheet)
        if not header_map:
            continue
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(v is None for v in row):
                continue
            rec: dict[str, Any] = {}
            for col_idx, field in header_map.items():
                if col_idx < len(row):
                    rec[field] = row[col_idx]
            all_rows.append(rec)
    return all_rows


def _parse_csv(path: Path) -> list[dict]:
    import csv
    with path.open(encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows: list[dict] = []
        for r in reader:
            norm: dict[str, Any] = {}
            for k, v in r.items():
                kl = (k or "").strip().lower()
                for field, aliases in NOA_COLUMN_ALIASES.items():
                    if any(a in kl for a in aliases):
                        norm[field] = v
                        break
            rows.append(norm)
        return rows


def _find_noa_header(sheet, max_scan: int = 12) -> tuple[int, dict[int, str]]:
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        mapping: dict[int, str] = {}
        for col_idx, cell in enumerate(cells):
            for field, aliases in NOA_COLUMN_ALIASES.items():
                if any(a in cell for a in aliases):
                    mapping[col_idx] = field
                    break
        if len(mapping) >= 2:
            return row_idx, mapping
    return 0, {}


def _has_substation_keyword(name: str, description: str) -> bool:
    blob = f"{name} {description}".lower()
    triggers = (
        "substation", "switching station", "sgt", "gsp",
        "gis", "ais", "transformer", "400/275", "400/132", "275/132",
        "bsp", "primary",
    )
    return any(t in blob for t in triggers)


def _noa_row_to_project(
    rec: dict,
    source_type: str,
    package_id: str | None,
    source_file: str,
) -> SubstationProject | None:
    from .ltds_excel_parser import _canonical_id, _safe_float
    from .riio_cv_extractor import _safe_int, _guess_upgrade, _guess_station_type

    name = (rec.get("scheme_name") or "").strip()
    if not name:
        return None

    to_raw = str(rec.get("to") or "").strip().upper()
    developer = _normalise_to(to_raw)
    if not developer:
        # NOA sometimes leaves TO blank for co-ordinated projects — try name match.
        developer = _guess_to_from_name(name) or "NGET"

    capex = _safe_float(rec.get("capex_gbp_m"))
    earliest = _safe_int(rec.get("earliest_ready"))
    voltage = _safe_float(rec.get("voltage_kv"))
    rec_text = str(rec.get("recommendation") or "").strip().lower()
    status = NOA_REC_STATUS.get(rec_text, "planned")
    noa_id = str(rec.get("noa_id") or "").strip() or f"{source_file}:{name}"
    region = str(rec.get("region") or "").strip() or _default_region_for(developer)

    descr = rec.get("description") or None
    upgrade = _guess_upgrade(name, descr or "", "")

    external_ids: dict[str, str] = {"noa": noa_id}
    if source_type == "CSNP":
        external_ids = {"csnp": noa_id}

    profile_url = None
    if package_id:
        profile_url = f"https://www.neso.energy/data/dataset/{package_id}"

    return SubstationProject(
        project_id=_canonical_id(developer, name, earliest),
        external_ids=external_ids,
        substation_name=name,
        station_type=_guess_station_type(name, descr or ""),
        parent_project=None,
        project_description=descr,
        upgrade_or_new=upgrade,
        region=region,
        lpa=None,
        county=None,
        country=COUNTRY_BY_DEVELOPER.get(developer, "England"),
        geom_wkt=None,
        osgb_easting=None,
        osgb_northing=None,
        parent_company=PARENT_COMPANY_BY_DEVELOPER.get(developer),
        developer=developer,
        voltage_kv_primary=voltage,
        voltage_kv_secondary=None,
        voltage_description=None,
        equipment_description=None,
        equipment_capacity_mva=None,
        construction_year=None,
        construction_date_detail=None,
        construction_status=status,
        operation_year=earliest,
        operation_date_detail=None,
        capex_gbp_millions=capex,
        capex_description=rec_text or None,
        capex_price_base_year=None,  # NOA uses its own price base (usually latest real terms).
        financial_description=rec_text or None,
        source_type=source_type,
        source_docket_id=noa_id,
        docket_profile_url=profile_url,
        source_links=[],
        capex_source_link=profile_url,
        last_updated=datetime.utcnow(),
        confidence=0.7 if rec_text in ("proceed", "proceed with plans") else 0.5,
        sme_reviewed=False,
    )


def _normalise_to(to: str) -> str:
    t = to.upper().replace(" ", "")
    if "NGET" in t or "NATIONALGRID" in t:   return "NGET"
    if "SPT" in t or "SPTRAN" in t:           return "SPT"
    if "SHE" in t or "SSEN-T" in t or "SSENT" in t: return "SHE-T"
    return t or ""


def _guess_to_from_name(name: str) -> str | None:
    n = name.lower()
    if any(k in n for k in ("scotland", "beauly", "denny", "dalmally", "peterhead", "carradale")):
        return "SHE-T"
    if any(k in n for k in ("border", "cumbernauld", "hunterston", "eccles", "torness")):
        return "SPT"
    return None


def _default_region_for(developer: str) -> str:
    if developer == "NGET":  return "NGET-England-Wales"
    if developer == "SPT":   return "SPT-Southern-Scotland"
    if developer == "SHE-T": return "SHE-T-Northern-Scotland"
    return developer
