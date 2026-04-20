"""RIIO ET2/ET3/ED2 business plan + cost matrix extractor.

RIIO (Revenue = Incentives + Innovation + Outputs) is Ofgem's price control
framework. Each TO/DNO publishes a multi-hundred-page business plan PDF plus
Cost Value ("CV") workbooks — these are the primary source of substation
CAPEX at project granularity.

Our job: parse the CV matrices (mostly tabular — capex by asset by year) and
pair each row with project narrative from the BP PDF to produce
SubstationProject records with the capex fields populated.

V1 status
---------
STUB — scaffolded with licensee list, file-naming conventions, and a
pluggable extraction interface. The actual heavy lifting (pdfplumber for
narrative, camelot for tables) is deferred because:
* The BP PDFs are 300-600 pages and cost table extraction is noisy — we
  need a representative sample pack to tune the regex + column heuristics.
* Several licensees published their CVs as .xlsx (not PDF); the same
  entry-point should handle both.

Target licensees
----------------
* Transmission (ET2 2021-26 + ET3 draft 2026-31):
    - NGET  — "National Grid Electricity Transmission" (England+Wales)
    - SPT   — "SP Transmission" (S Scotland)
    - SHE-T — "Scottish Hydro Electric Transmission" / SSEN-T (N Scotland)
* Distribution (ED2 2023-28):
    - UKPN, SSEN, SPEN, NGED, NPG, ENWL

Each licensee's CV exposes asset-level capex by year. We key on "Substation"
and "Switching Station" rows and pull:
    * project name (column varies — typically "Scheme Name" or "Asset Name")
    * project description
    * capex total (£m, 18/19 base for ET2, 20/21 base for ED2)
    * start/end years
    * driver type (Load-related / Non-load / Replacement / New connection)

See Ofgem's RIIO library:
    https://www.ofgem.gov.uk/energy-policy-and-regulation/policy-and-regulatory-programmes/network-price-controls-riio-2
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from .schema import (
    SubstationProject,
    PARENT_COMPANY_BY_DEVELOPER,
    COUNTRY_BY_DEVELOPER,
    RIIO_PRICE_BASE_BY_DEVELOPER,
)

log = logging.getLogger("princeps.substation_tracker.riio_cv")


# ---------------------------------------------------------------------------
# File classification
# ---------------------------------------------------------------------------
RIIO_ET_LICENSEES = ("NGET", "SPT", "SHE-T")
RIIO_ED_LICENSEES = ("UKPN", "SSEN", "SPEN", "NGED", "NPG", "ENWL")

# Heuristic: file name stem hints at source_type.
_ET_MARKERS = ("et2", "et3", "transmission")
_ED_MARKERS = ("ed2", "ed3", "distribution")

# Heuristic CV column aliases — we do a case-insensitive contains check.
CV_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "scheme_name":   ("scheme name", "project name", "asset name", "site name", "substation name"),
    "description":   ("description", "scope", "scheme description"),
    "capex_gbp_m":   ("total capex", "capex £m", "capex_gbp_m", "total cost £m", "cost £m"),
    "start_year":    ("start year", "start", "construction start"),
    "end_year":      ("end year", "end", "commissioning", "energisation"),
    "driver":        ("driver", "cost category", "cost type", "workstream"),
    "region":        ("region", "licence area", "zone", "location"),
    "voltage_kv":    ("voltage kv", "voltage", "kv"),
    "mva":           ("mva", "rating mva", "capacity mva"),
}


def classify_source(path: Path, developer: str) -> str:
    """Return the SourceType literal ('RIIO_ET' or 'RIIO_ED') for a file."""
    name = path.name.lower()
    if developer.upper() in RIIO_ET_LICENSEES:
        return "RIIO_ET"
    if developer.upper() in RIIO_ED_LICENSEES:
        return "RIIO_ED"
    if any(m in name for m in _ET_MARKERS):
        return "RIIO_ET"
    if any(m in name for m in _ED_MARKERS):
        return "RIIO_ED"
    # Default to ED — the majority of substation-level capex is distribution.
    return "RIIO_ED"


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def extract_from_riio_cv(
    path: Path,
    developer: str,
    *,
    price_base_year: int | None = None,
    sheet_name: str | None = None,
) -> list[SubstationProject]:
    """Extract substation CAPEX rows from a RIIO CV file.

    Accepts both .xlsx (preferred) and .pdf. The extraction logic differs:
    * .xlsx → use openpyxl to find substation rows, map columns via
      `CV_COLUMN_ALIASES`, emit SubstationProject records.
    * .pdf  → use pdfplumber for narrative + camelot for tables; match
      narrative → table rows by scheme name.

    Args:
        path: CV workbook or PDF.
        developer: Licensee code (NGET/SPT/SHE-T/UKPN/...).
        price_base_year: Override the default for this licensee.
        sheet_name: If .xlsx and multi-sheet, pin the CV matrix sheet.

    Returns:
        List of SubstationProject records (empty if nothing matches).
    """
    if not path.exists():
        log.warning("RIIO CV file does not exist: %s", path)
        return []

    source_type = classify_source(path, developer)
    price_base = price_base_year or RIIO_PRICE_BASE_BY_DEVELOPER.get(developer.upper())

    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _extract_xlsx(path, developer, source_type, price_base, sheet_name)
    if suffix == ".pdf":
        return _extract_pdf(path, developer, source_type, price_base)
    log.warning("Unsupported RIIO CV file type: %s", path.suffix)
    return []


# ---------------------------------------------------------------------------
# XLSX path
# ---------------------------------------------------------------------------
def _extract_xlsx(
    path: Path,
    developer: str,
    source_type: str,
    price_base: int | None,
    sheet_name: str | None,
) -> list[SubstationProject]:
    """Extract substation rows from a CV workbook via openpyxl.

    We scan every sheet (or the named sheet) for a header row that contains
    at least one capex-like column. For each data row that mentions
    'substation' or 'switching station' in the scheme/description, emit a
    SubstationProject.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        log.error("openpyxl is required for .xlsx RIIO CVs — install in .venv")
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [wb[sheet_name]] if sheet_name else list(wb.worksheets)
    out: list[SubstationProject] = []

    for sheet in sheets:
        header_row, header_map = _find_cv_header(sheet)
        if not header_map:
            continue
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(v is None for v in row):
                continue
            rec = _apply_header_map(row, header_map)
            name = (rec.get("scheme_name") or "").strip() if rec.get("scheme_name") else ""
            descr = (rec.get("description") or "").lower() if rec.get("description") else ""
            if not name or not _is_substation_row(name, descr):
                continue
            proj = _row_to_project(
                rec=rec,
                developer=developer,
                source_type=source_type,
                price_base=price_base,
                source_file=path.name,
            )
            if proj:
                out.append(proj)
    log.info("RIIO CV %s: extracted %d substation rows", path.name, len(out))
    return out


def _find_cv_header(sheet, max_scan: int = 15) -> tuple[int, dict[int, str]]:
    """Locate the header row by looking for at least two alias matches."""
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        mapping: dict[int, str] = {}
        for col_idx, cell in enumerate(cells):
            for field, aliases in CV_COLUMN_ALIASES.items():
                if any(a in cell for a in aliases):
                    mapping[col_idx] = field
                    break
        if len(mapping) >= 2:
            return row_idx, mapping
    return 0, {}


def _apply_header_map(row: tuple, header_map: dict[int, str]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for col_idx, field in header_map.items():
        if col_idx < len(row):
            out[field] = row[col_idx]
    return out


def _is_substation_row(name: str, description: str) -> bool:
    n = name.lower()
    d = description.lower()
    triggers = ("substation", "switching station", "gsp", "bsp", "primary", "132/33", "400/132", "275/132", "sgt")
    return any(t in n or t in d for t in triggers)


def _row_to_project(
    rec: dict,
    developer: str,
    source_type: str,
    price_base: int | None,
    source_file: str,
) -> SubstationProject | None:
    from .ltds_excel_parser import _canonical_id, _safe_float  # reuse helpers

    name = (rec.get("scheme_name") or "").strip()
    if not name:
        return None
    capex = _safe_float(rec.get("capex_gbp_m"))
    start = _safe_int(rec.get("start_year"))
    end = _safe_int(rec.get("end_year"))
    voltage = _safe_float(rec.get("voltage_kv"))
    mva = _safe_float(rec.get("mva"))
    region = str(rec.get("region") or "").strip() or _default_region_for(developer)
    driver = str(rec.get("driver") or "").strip()

    descr = rec.get("description") or None
    upgrade = _guess_upgrade(name, descr or "", driver)

    return SubstationProject(
        project_id=_canonical_id(developer, name, end or start),
        external_ids={"riio_cv": f"{developer}:{source_file}:{name}"},
        substation_name=name,
        station_type=_guess_station_type(name, descr or ""),
        parent_project=None,
        project_description=descr,
        upgrade_or_new=upgrade,
        region=region,
        lpa=None,
        county=None,
        country=COUNTRY_BY_DEVELOPER.get(developer.upper(), "England"),
        geom_wkt=None,
        osgb_easting=None,
        osgb_northing=None,
        parent_company=PARENT_COMPANY_BY_DEVELOPER.get(developer.upper()),
        developer=developer.upper(),
        voltage_kv_primary=voltage,
        voltage_kv_secondary=None,
        voltage_description=None,
        equipment_description=None,
        equipment_capacity_mva=mva,
        construction_year=start,
        construction_date_detail=None,
        construction_status="planned" if (start and start > datetime.utcnow().year) else "under_construction",
        operation_year=end,
        operation_date_detail=None,
        capex_gbp_millions=capex,
        capex_description=driver or None,
        capex_price_base_year=price_base,
        financial_description=driver or None,
        source_type=source_type,
        source_docket_id=f"{developer}:{source_file}",
        docket_profile_url=None,
        source_links=[],
        capex_source_link=None,
        last_updated=datetime.utcnow(),
        confidence=0.65,  # CV is authoritative for capex but noisy for project scope
        sme_reviewed=False,
    )


# ---------------------------------------------------------------------------
# PDF path — STUB
# ---------------------------------------------------------------------------
def _extract_pdf(
    path: Path,
    developer: str,
    source_type: str,
    price_base: int | None,
) -> list[SubstationProject]:
    """Extract via pdfplumber (narrative) + camelot-py (tables).

    STUB — see module docstring. This is intentionally left empty until we
    have representative RIIO BP PDFs to tune against. The flow when wired:

        import pdfplumber
        with pdfplumber.open(path) as pdf:
            narrative_blocks = _find_substation_narrative(pdf)
        import camelot
        tables = camelot.read_pdf(str(path), pages="all", flavor="lattice")
        cv_rows = _match_narrative_to_table_rows(narrative_blocks, tables)
        return [_row_to_project(...) for row in cv_rows]

    For now we log and return [] so the pipeline degrades gracefully.
    """
    log.warning("RIIO CV PDF extraction not yet implemented (%s) — returning empty", path.name)
    return []


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _safe_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        m = re.search(r"\d{4}", str(v))
        return int(m.group(0)) if m else None


def _guess_upgrade(name: str, description: str, driver: str) -> str:
    blob = " ".join([name, description, driver]).lower()
    if any(k in blob for k in ("new build", "new substation", "greenfield")):
        return "new"
    if any(k in blob for k in ("replace", "replacement", "asset renewal", "end of life", "eol")):
        return "replacement"
    if any(k in blob for k in ("upgrade", "reinforcement", "reinforce", "uprate", "extension", "augment")):
        return "upgrade"
    return "unknown"


def _guess_station_type(name: str, description: str) -> str:
    blob = " ".join([name, description]).lower()
    if "switching station" in blob or "switching" in blob and "station" in blob:
        return "switching_station"
    return "substation"


def _default_region_for(developer: str) -> str:
    dev = developer.upper()
    if dev == "NGET":  return "NGET"
    if dev == "SPT":   return "SPT"
    if dev == "SHE-T": return "SHE-T"
    return dev
