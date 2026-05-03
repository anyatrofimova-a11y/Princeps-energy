"""LTDS Excel workbook parser → canonical SubstationProject list.

Each of the 6 UK DNOs publishes a Long Term Development Statement every
November. The file is an Excel workbook, but the schema drifts between DNOs
(different sheet names, different column orders, different units). We handle
this with a dispatcher: each DNO has a dedicated column-mapping callable that
returns a list of `SubstationProject` instances.

V1 status
---------
* UKPN — **wired end-to-end** (HTTP fetch via OpenDataSoft, parse, return).
  UKPN publishes its LTDS primary/grid substations dataset as CSV/JSON on
  the OpenDataSoft portal (dataset ids `ltds-tab-6-grid-substations-loading`
  and `ltds-tab-2-primary-substations-loading`). This parser handles both
  the downloaded .xlsx form and the ODS JSON form.
* SSEN, SPEN, NPG, ENWL — **v1 wired** with candidate column mappings based
  on the DNOs' published LTDS explanatory notes. Each parser auto-detects
  the header row, falls back across a list of plausible header aliases for
  each field, and tolerates missing columns with warnings rather than
  hard-failing. Confidence flags are attached per DNO — see the docstring
  on each `parse_*_workbook` function.
* NGED — **stubbed** (uses CIM RDF/XML, not Excel). Canonical NGED path is
  `utils/ltds_cim_ingester.py`.

Each DNO-specific adapter is registered in DNO_DISPATCHERS; `parse_ltds_workbook`
looks up the right one by the `dno` argument.

CLI usage
---------
    python -m utils.substation_tracker.ltds_excel_parser \
        --dno ssen --file data/ltds/ssen/2025.xlsx --dry-run

    python -m utils.substation_tracker.ltds_excel_parser \
        --dno npg --file data/ltds/npg/NPG_LTDS_2024.xlsx \
        --db-url "$DATABASE_URL"

When `--db-url` is omitted or `--dry-run` is passed the parser only emits
parsed records to stdout (JSON) — nothing touches the DB.

TODO after v1
-------------
* Replace candidate header lists with verified column names once real
  workbook samples for each DNO are in `data/ltds/<dno>/`. Any header
  outside the candidate list will currently fall through to `None`.
"""

from __future__ import annotations

import io
import json
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping, Sequence

from .schema import (
    SubstationProject,
    PARENT_COMPANY_BY_DEVELOPER,
    COUNTRY_BY_DEVELOPER,
)

log = logging.getLogger("princeps.substation_tracker.ltds_excel")

# ---------------------------------------------------------------------------
# Dependency guards — openpyxl is a heavy import; we only need it when we
# actually parse a workbook. httpx is pulled in elsewhere in the project so
# it's safe.
# ---------------------------------------------------------------------------
def _load_openpyxl():
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for LTDS Excel parsing — "
            "install with `.venv/bin/pip install openpyxl`"
        ) from exc


# ---------------------------------------------------------------------------
# UKPN — primary exemplar (end-to-end)
# ---------------------------------------------------------------------------
UKPN_ODS_BASE = "https://ukpowernetworks.opendatasoft.com/api/explore/v2.1/catalog/datasets"
UKPN_ODS_PRIMARY_DS = "ukpn-primary-postcode-area"  # primary substation loading info
UKPN_ODS_GRID_DS = "ukpn-grid-and-primary-sites"    # grid+primary sites w/ coords

# UKPN licence area codes → region label.
UKPN_LICENCE_AREAS = {
    "EPN": "Eastern Power Networks",
    "SPN": "South Eastern Power Networks",
    "LPN": "London Power Networks",
}


async def fetch_ukpn_ods_records(dataset_id: str = UKPN_ODS_GRID_DS, limit: int = 100) -> list[dict]:
    """Fetch all UKPN LTDS records via the OpenDataSoft v2.1 catalog API.

    We reuse the project's shared UKPN ODS endpoint. No auth required for
    public datasets, but an `apikey` is honoured if `UKPN_ODS_APIKEY` is set.
    """
    import os
    import httpx

    apikey = os.environ.get("UKPN_ODS_APIKEY", "").strip()
    auth = f"&apikey={apikey}" if apikey else ""
    records: list[dict] = []
    offset = 0
    async with httpx.AsyncClient(timeout=30) as client:
        while True:
            url = f"{UKPN_ODS_BASE}/{dataset_id}/records?limit={limit}&offset={offset}{auth}"
            r = await client.get(url)
            if r.status_code != 200:
                log.warning("UKPN ODS %s HTTP %d at offset %d", dataset_id, r.status_code, offset)
                break
            data = r.json()
            batch = data.get("results", [])
            if not batch:
                break
            records.extend(batch)
            total = int(data.get("total_count", 0) or 0)
            offset += limit
            if offset >= total:
                break
    log.info("UKPN ODS %s: fetched %d records", dataset_id, len(records))
    return records


def _ukpn_to_project(rec: dict) -> SubstationProject | None:
    """Map one UKPN ODS record → SubstationProject. Returns None if unusable."""
    name = (rec.get("sitename") or rec.get("sitefunctionallocation") or "").strip()
    if not name:
        return None

    licence_area = (rec.get("licencearea") or "").strip()
    region = UKPN_LICENCE_AREAS.get(licence_area, licence_area or "Unknown")

    # Voltage — UKPN exposes single voltage for the site, so use primary only.
    voltage = _safe_float(rec.get("sitevoltage"))

    # Transformer rating — max of summer/winter as capacity proxy.
    trans_summer = _safe_float(rec.get("transratingsummer"))
    trans_winter = _safe_float(rec.get("transratingwinter"))
    capacity_mva: float | None = None
    if trans_summer or trans_winter:
        capacity_mva = max(x for x in (trans_summer, trans_winter) if x is not None)

    # Coords — prefer spatial_coordinates, fallback to geo_point_2d.
    lat, lon = _extract_latlon(rec)
    wkt = f"POINT({lon} {lat})" if lat and lon else None

    # OSGB is not provided by UKPN ODS — pipeline layer will convert from WGS84.
    project_id = _canonical_id("UKPN", name, None)

    # Site type → station_type + upgrade/new guess.
    site_type = (rec.get("sitetype") or "").lower()
    station_type = "substation"  # UKPN grid/primary are all substations, no switching stations in this dataset.

    return SubstationProject(
        project_id=project_id,
        external_ids={"ltds": rec.get("sitefunctionallocation") or name},
        substation_name=name,
        station_type=station_type,
        parent_project=None,
        project_description=site_type or None,
        upgrade_or_new="unknown",  # LTDS is a snapshot — upgrade classification lives in RIIO CV.
        region=region,
        lpa=None,
        county=rec.get("county"),
        country=COUNTRY_BY_DEVELOPER.get("UKPN", "England"),
        geom_wkt=wkt,
        osgb_easting=None,
        osgb_northing=None,
        parent_company=PARENT_COMPANY_BY_DEVELOPER.get("UKPN"),
        developer="UKPN",
        voltage_kv_primary=voltage,
        voltage_kv_secondary=None,
        voltage_description=None,
        equipment_description=rec.get("sitetype"),
        equipment_capacity_mva=capacity_mva,
        construction_year=None,
        construction_date_detail=None,
        construction_status="complete",  # LTDS only lists operational assets.
        operation_year=None,
        operation_date_detail=None,
        capex_gbp_millions=None,
        capex_description=None,
        capex_price_base_year=None,
        financial_description=None,
        source_type="LTDS",
        source_docket_id=rec.get("sitefunctionallocation"),
        docket_profile_url=f"https://ukpowernetworks.opendatasoft.com/explore/dataset/{UKPN_ODS_GRID_DS}/",
        source_links=[],
        capex_source_link=None,
        last_updated=datetime.utcnow(),
        confidence=0.85,  # UKPN ODS is authoritative for UKPN assets.
        sme_reviewed=False,
    )


def parse_ukpn_workbook(path: Path | None = None, records: Iterable[dict] | None = None) -> list[SubstationProject]:
    """Parse UKPN LTDS into SubstationProject records.

    Accepts either a local xlsx path OR pre-fetched ODS records. When only a
    path is given, we read the workbook directly; when `records` are given
    (from `fetch_ukpn_ods_records`) we skip the workbook.
    """
    if records is not None:
        out = [p for p in (_ukpn_to_project(r) for r in records) if p is not None]
        log.info("UKPN LTDS: produced %d SubstationProject records from %d input", len(out), sum(1 for _ in records))
        return out

    if path is None:
        return []

    openpyxl = _load_openpyxl()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[SubstationProject] = []
    # Heuristic: iterate all sheets, look for a header row containing
    # 'Site Name' or similar, then read records.
    for sheet in wb.worksheets:
        header_row, headers = _find_header_row(sheet, candidates=("site name", "sitename", "substation"))
        if not headers:
            continue
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(v is None for v in row):
                continue
            rec = {h: row[i] if i < len(row) else None for i, h in enumerate(headers)}
            # Normalise ODS-style keys so `_ukpn_to_project` can reuse.
            norm = {
                "sitename": rec.get("site name") or rec.get("sitename") or rec.get("substation name"),
                "sitefunctionallocation": rec.get("functional location") or rec.get("site functional location"),
                "licencearea": rec.get("licence area") or rec.get("licence_area"),
                "sitevoltage": rec.get("site voltage (kv)") or rec.get("voltage kv") or rec.get("voltage"),
                "transratingsummer": rec.get("trans rating summer (mva)") or rec.get("summer rating mva"),
                "transratingwinter": rec.get("trans rating winter (mva)") or rec.get("winter rating mva"),
                "sitetype": rec.get("site type") or rec.get("sitetype"),
                "county": rec.get("county"),
            }
            proj = _ukpn_to_project(norm)
            if proj:
                rows.append(proj)
    log.info("UKPN LTDS workbook: produced %d rows from %s", len(rows), path)
    return rows


# ---------------------------------------------------------------------------
# Generic Excel header-matching engine used by SSEN/SPEN/NPG/ENWL
#
# Each DNO publishes a workbook with mildly drifting column headers year on
# year. Rather than hard-code one header string per field, we give the
# engine an ordered list of candidate aliases; the first one present in
# the sheet's header row wins. This means if a DNO renames
# "Firm Capacity (MVA)" → "Firm Rating (MVA)" next year we only add one
# alias to the config, no dispatcher code changes.
# ---------------------------------------------------------------------------

# Canonical field name → ordered list of header aliases we will accept.
# Matching is case-insensitive with whitespace collapsed. Extend rather
# than replace entries when verifying against real workbooks.
FieldAliases = Mapping[str, Sequence[str]]


# Best-effort Easting/Northing → WGS84 converter. pyproj is a hard dep of
# the project (see requirements.txt), but if it isn't importable we degrade
# to returning None so the parser still runs in limited environments.
def _osgb_to_wgs84(easting: float | None, northing: float | None) -> tuple[float | None, float | None]:
    """OSGB36 BNG (EPSG:27700) → WGS84 (EPSG:4326). Returns (lat, lon)."""
    if easting is None or northing is None:
        return None, None
    try:
        from pyproj import Transformer  # type: ignore
    except ImportError:
        log.warning("pyproj not available — OSGB coords passed through without transform")
        return None, None
    try:
        t = Transformer.from_crs("EPSG:27700", "EPSG:4326", always_xy=True)
        lon, lat = t.transform(float(easting), float(northing))
        # pyproj returns NaN for out-of-bounds input — guard.
        if lat != lat or lon != lon:  # NaN check
            return None, None
        return float(lat), float(lon)
    except Exception as exc:
        log.debug("pyproj transform failed for E=%s N=%s: %s", easting, northing, exc)
        return None, None


def _norm_header(h: Any) -> str:
    """Normalise a sheet header for alias matching."""
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h).strip()).lower()


def _resolve_aliases(headers: list[str], aliases: FieldAliases) -> dict[str, int | None]:
    """For each canonical field, return the column index of the first alias
    that appears in `headers`. Unresolved fields map to None.
    """
    norm = [_norm_header(h) for h in headers]
    resolved: dict[str, int | None] = {}
    for field, candidates in aliases.items():
        idx: int | None = None
        for cand in candidates:
            c = _norm_header(cand)
            if c in norm:
                idx = norm.index(c)
                break
            # Try substring match for verbose column headers like
            # "Forecast Winter Peak Demand (MW) 2028/29".
            for i, h in enumerate(norm):
                if c and c in h:
                    idx = i
                    break
            if idx is not None:
                break
        resolved[field] = idx
    return resolved


def _parse_workbook_pandas(
    path: Path,
    *,
    sheet_aliases: Sequence[str],
    field_aliases: FieldAliases,
    header_scan_rows: int = 10,
) -> list[dict[str, Any]]:
    """Generic pandas-based LTDS sheet extractor.

    Looks for the first sheet whose name matches any alias in `sheet_aliases`
    (case-insensitive substring), scans the top `header_scan_rows` for the
    first row containing the 'name' field's first alias, then reads from
    that row onwards.

    Returns a list of raw-field-keyed dicts (one per data row); the
    DNO-specific adapter then maps these to SubstationProject records.
    """
    import pandas as pd

    # Load all sheet names first so we can pick.
    try:
        xl = pd.ExcelFile(path, engine="openpyxl")
    except Exception as exc:
        log.error("Failed to open workbook %s: %s", path, exc)
        return []

    picked_sheets: list[str] = []
    for s in xl.sheet_names:
        ns = _norm_header(s)
        for alias in sheet_aliases:
            if _norm_header(alias) in ns:
                picked_sheets.append(s)
                break
    if not picked_sheets:
        # Fallback: take all sheets — some workbooks concentrate everything
        # in one tab named "Substations" / "Loading" and we don't want to
        # miss it because of a novel sheet title.
        picked_sheets = list(xl.sheet_names)
        log.info("LTDS %s: no sheets matched %s — scanning all %d sheets",
                 path.name, list(sheet_aliases), len(picked_sheets))

    rows_out: list[dict[str, Any]] = []
    name_aliases = [_norm_header(a) for a in field_aliases.get("name", ())]

    for sheet in picked_sheets:
        # Read without header to scan for the real header row.
        try:
            df_raw = xl.parse(sheet, header=None, dtype=object)
        except Exception as exc:
            log.warning("LTDS %s: failed to parse sheet %s: %s", path.name, sheet, exc)
            continue
        if df_raw.empty:
            continue

        header_row_idx: int | None = None
        for ridx in range(min(header_scan_rows, len(df_raw))):
            row_vals = [_norm_header(v) for v in df_raw.iloc[ridx].tolist()]
            # Header row wins if it contains any of the 'name' aliases — or
            # any substring thereof.
            if any(
                any(a and (a == rv or a in rv) for rv in row_vals)
                for a in name_aliases
            ):
                header_row_idx = ridx
                break
        if header_row_idx is None:
            log.debug("LTDS %s/%s: no header row found", path.name, sheet)
            continue

        headers = [str(h) if h is not None else "" for h in df_raw.iloc[header_row_idx].tolist()]
        col_index = _resolve_aliases(headers, field_aliases)

        missing = [k for k, v in col_index.items() if v is None]
        if missing:
            log.info("LTDS %s/%s: columns not matched -> %s (header row: %s)",
                     path.name, sheet, missing, [h for h in headers if h])

        data = df_raw.iloc[header_row_idx + 1:].reset_index(drop=True)
        for _, drow in data.iterrows():
            rec: dict[str, Any] = {}
            for field, idx in col_index.items():
                if idx is None:
                    rec[field] = None
                else:
                    val = drow.iloc[idx] if idx < len(drow) else None
                    if val is None or (isinstance(val, float) and val != val):  # NaN
                        rec[field] = None
                    else:
                        rec[field] = val
            rec["_sheet"] = sheet
            rec["_source_path"] = str(path)
            # Skip fully empty data rows.
            if rec.get("name") in (None, "", "nan"):
                continue
            rows_out.append(rec)

    log.info("LTDS %s: extracted %d rows from %d sheet(s)",
             path.name, len(rows_out), len(picked_sheets))
    return rows_out


def _raw_row_to_project(
    rec: dict[str, Any],
    *,
    developer: str,
    default_region: str,
    default_country: str,
    docket_profile_url: str | None,
) -> SubstationProject | None:
    """Map a raw-header-keyed dict → SubstationProject.

    Expects the following canonical keys in `rec` (any may be missing):
        name, voltage_kv, firm_capacity_mva, demand_headroom_mw,
        peak_demand_mw, easting, northing, licence_area, site_type,
        external_id, operation_year
    """
    name = rec.get("name")
    if not name or str(name).strip().lower() in ("nan", "none", ""):
        return None
    name_s = str(name).strip()

    voltage = _safe_float(rec.get("voltage_kv"))
    capacity_mva = _safe_float(rec.get("firm_capacity_mva"))
    headroom_mw = _safe_float(rec.get("demand_headroom_mw"))
    peak_demand = _safe_float(rec.get("peak_demand_mw"))
    easting = _safe_float(rec.get("easting"))
    northing = _safe_float(rec.get("northing"))
    licence_area = (rec.get("licence_area") or "").strip() if rec.get("licence_area") else ""
    region = licence_area or default_region
    site_type = (rec.get("site_type") or "").strip() if rec.get("site_type") else None
    op_year = None
    if rec.get("operation_year") is not None:
        try:
            op_year = int(float(str(rec.get("operation_year")).strip()))
        except (ValueError, TypeError):
            op_year = None

    lat, lon = _osgb_to_wgs84(easting, northing)
    wkt = f"POINT({lon} {lat})" if (lat is not None and lon is not None) else None

    external_id = rec.get("external_id") or name_s
    project_id = _canonical_id(developer, name_s, op_year)

    # Build a financial_description that carries headroom so downstream
    # consumers (grid_substations upsert) can pull it out.
    fin_bits = []
    if peak_demand is not None:
        fin_bits.append(f"peak demand {peak_demand:.1f} MW")
    if headroom_mw is not None:
        fin_bits.append(f"demand headroom {headroom_mw:.1f} MW")
    financial_description = "; ".join(fin_bits) or None

    # Country mapping: SSEN SHEPD = Scotland; SPEN SPD = Scotland, SPM =
    # Wales/England. The parent parser passes the correct `default_country`.
    country = default_country  # type: ignore[assignment]

    return SubstationProject(
        project_id=project_id,
        external_ids={"ltds": str(external_id)},
        substation_name=name_s,
        station_type="substation",
        parent_project=None,
        project_description=site_type,
        upgrade_or_new="unknown",
        region=region,
        lpa=None,
        county=None,
        country=country,  # type: ignore[arg-type]
        geom_wkt=wkt,
        osgb_easting=easting,
        osgb_northing=northing,
        parent_company=PARENT_COMPANY_BY_DEVELOPER.get(developer),
        developer=developer,
        voltage_kv_primary=voltage,
        voltage_kv_secondary=None,
        voltage_description=None,
        equipment_description=site_type,
        equipment_capacity_mva=capacity_mva,
        construction_year=None,
        construction_date_detail=None,
        construction_status="complete",
        operation_year=op_year,
        operation_date_detail=None,
        capex_gbp_millions=None,
        capex_description=None,
        capex_price_base_year=None,
        financial_description=financial_description,
        source_type="LTDS",
        source_docket_id=str(external_id) if external_id else None,
        docket_profile_url=docket_profile_url,
        source_links=[],
        capex_source_link=None,
        last_updated=datetime.utcnow(),
        confidence=0.7,  # slightly lower than UKPN (0.85) — aliases may drift
        sme_reviewed=False,
    )


# ---------------------------------------------------------------------------
# SSEN — Scottish & Southern Electricity Networks
# ---------------------------------------------------------------------------
# Confidence: MEDIUM. Landing page known; column headers are the "Site
# Name / Nominal Voltage / Firm Capacity" pattern common to the SHEPD
# and SEPD workbooks published under SSEN's LTDS page. Aliases here cover
# the most-likely names; extend once a sample workbook lands in
# `data/ltds/ssen/`.
SSEN_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "name": (
        "Site Name", "Substation Name", "Substation", "Name",
    ),
    "voltage_kv": (
        "Nominal Voltage (kV)", "Voltage (kV)", "Voltage [kV]", "Voltage",
        "Operating Voltage (kV)",
    ),
    "firm_capacity_mva": (
        "Firm Capacity (MVA)", "Firm Capacity MVA", "Firm Rating (MVA)",
        "Firm Capacity",
    ),
    "demand_headroom_mw": (
        "Demand Headroom (MW)", "Headroom (MW)", "Demand HR (MW)",
        "Available Demand Headroom (MW)", "Headroom MW",
    ),
    "peak_demand_mw": (
        "Forecast Winter Peak Demand (MW)", "Winter Peak Demand (MW)",
        "Peak Demand (MW)", "Maximum Demand (MW)",
    ),
    "easting": (
        "OS Easting", "Easting", "Grid Easting", "Easting (m)",
    ),
    "northing": (
        "OS Northing", "Northing", "Grid Northing", "Northing (m)",
    ),
    "licence_area": (
        "Licence Area", "Region", "Zone", "Licence",
    ),
    "site_type": (
        "Primary / Grid", "Site Type", "Type", "Substation Type", "Category",
    ),
    "external_id": (
        "Site Code", "Site ID", "Substation Code", "Asset ID", "ID", "SAP ID",
    ),
    "operation_year": (
        "Commissioning Year", "Year Commissioned", "Commission Year",
    ),
}

SSEN_SHEET_ALIASES: tuple[str, ...] = (
    "Primary Substations", "Grid Substations", "BSP", "Substation",
    "Primary", "Grid", "Network Capacity",
)


def parse_ssen_workbook(path: Path | None = None, **_: Any) -> list[SubstationProject]:
    """SSEN LTDS parser.

    SSEN publishes two workbooks under its LTDS page: SEPD (Southern —
    England) and SHEPD (Scottish Hydro — Scotland). Licence area and
    country are inferred from the filename; the parser falls back to
    "SSEN" / England when the filename is ambiguous.

    TODO — column-mapping (candidate aliases):
      * Name            : 'Site Name' | 'Substation Name'
      * Voltage         : 'Nominal Voltage (kV)' | 'Voltage (kV)'
      * Firm capacity   : 'Firm Capacity (MVA)'
      * Demand headroom : 'Demand Headroom (MW)' | 'Headroom (MW)'
      * Coordinates     : 'OS Easting' + 'OS Northing' (OSGB 27700)
      * Licence area    : 'Licence Area' ('SEPD'|'SHEPD')

    When a real 2025 workbook is in place, verify these aliases against
    the actual header row and add any new spellings to
    `SSEN_FIELD_ALIASES`.
    """
    if path is None:
        log.warning("SSEN LTDS: no path supplied — returning empty")
        return []
    path = Path(path)
    if not path.exists():
        log.warning("SSEN LTDS: %s not found — returning empty", path)
        return []

    _load_openpyxl()  # dependency guard — surfaces a clear error

    name = path.name.lower()
    if "shepd" in name or "hepd" in name or "hydro" in name:
        licence_area = "SHEPD"
        default_country = "Scotland"
    elif "sepd" in name or "southern" in name:
        licence_area = "SEPD"
        default_country = "England"
    else:
        licence_area = "SSEN"
        default_country = "England"

    raw_rows = _parse_workbook_pandas(
        path,
        sheet_aliases=SSEN_SHEET_ALIASES,
        field_aliases=SSEN_FIELD_ALIASES,
    )

    projects: list[SubstationProject] = []
    for rec in raw_rows:
        # If the workbook doesn't have a licence area column, fall back to
        # the filename-derived value.
        if not rec.get("licence_area"):
            rec["licence_area"] = licence_area
        proj = _raw_row_to_project(
            rec,
            developer="SSEN",
            default_region=licence_area,
            default_country=default_country,
            docket_profile_url="https://www.ssen.co.uk/our-services/"
                               "industry-data-and-links/long-term-development-statement/",
        )
        if proj:
            projects.append(proj)
    log.info("SSEN LTDS: %d projects parsed from %s", len(projects), path.name)
    return projects


# ---------------------------------------------------------------------------
# SPEN — SP Energy Networks
# ---------------------------------------------------------------------------
# Confidence: MEDIUM. Two licence areas (SPD = Scotland, SPM = Manweb NW
# England + N Wales) typically ship in a single workbook with a
# licence-area column. Column aliases cover both "[kV]" and "(kV)" brackets
# (SPEN has historically used the square-bracket form).
SPEN_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "name": (
        "Name", "Substation Name", "Site Name", "BSP Name",
    ),
    "voltage_kv": (
        "Voltage [kV]", "Voltage (kV)", "Nominal Voltage (kV)", "Voltage kV",
    ),
    "firm_capacity_mva": (
        "Transformer MVA", "Firm Capacity (MVA)", "Firm Capacity MVA",
        "Rating (MVA)", "Substation Firm Capacity (MVA)",
    ),
    "demand_headroom_mw": (
        "Demand Headroom (MW)", "Headroom (MW)", "Demand HR (MW)",
        "Spare Capacity (MW)",
    ),
    "peak_demand_mw": (
        "Forecast Demand 2028 (MW)", "Forecast Demand (MW)",
        "Peak Demand (MW)", "Maximum Demand (MW)",
    ),
    "easting": (
        "Easting", "OS Easting", "X", "BNG Easting",
    ),
    "northing": (
        "Northing", "OS Northing", "Y", "BNG Northing",
    ),
    "licence_area": (
        "Licence Area", "Licence", "SPD/SPM", "Distribution Area",
    ),
    "site_type": (
        "Type", "Substation Type", "BSP/Primary", "Site Type",
    ),
    "external_id": (
        "Site Code", "Substation Code", "BSP Code", "Asset ID", "ID",
    ),
    "operation_year": (
        "Commissioning Year", "Year Commissioned",
    ),
}

SPEN_SHEET_ALIASES: tuple[str, ...] = (
    "Demand Substations", "BSP", "Primary", "Grid",
    "Substation", "Network Capacity", "Loading",
)


def parse_spen_workbook(path: Path | None = None, **_: Any) -> list[SubstationProject]:
    """SPEN (SP Energy Networks) LTDS parser.

    SPEN has two licence areas — SPD (Scotland) and SPM (Manweb — NW
    England + N Wales). Both typically ship in a single workbook with a
    licence-area column; the parser uses that column if present, falling
    back to a filename-derived value.

    TODO — column-mapping (candidate aliases):
      * Name            : 'Name' | 'Substation Name'
      * Voltage         : 'Voltage [kV]' | 'Voltage (kV)'
      * Firm capacity   : 'Transformer MVA' | 'Firm Capacity (MVA)'
      * Demand headroom : 'Demand Headroom (MW)' | 'Headroom (MW)'
      * Coordinates     : 'Easting' + 'Northing' (OSGB 27700)
      * Licence area    : 'Licence Area' ('SPD'|'SPM')
    """
    if path is None:
        log.warning("SPEN LTDS: no path supplied — returning empty")
        return []
    path = Path(path)
    if not path.exists():
        log.warning("SPEN LTDS: %s not found — returning empty", path)
        return []

    _load_openpyxl()

    name = path.name.lower()
    if "spm" in name or "manweb" in name:
        fallback_area = "SPM"
        fallback_country = "Wales"  # mostly — adapter could be sharpened later
    elif "spd" in name or "scot" in name:
        fallback_area = "SPD"
        fallback_country = "Scotland"
    else:
        fallback_area = "SPEN"
        fallback_country = "Scotland"

    raw_rows = _parse_workbook_pandas(
        path,
        sheet_aliases=SPEN_SHEET_ALIASES,
        field_aliases=SPEN_FIELD_ALIASES,
    )

    projects: list[SubstationProject] = []
    for rec in raw_rows:
        la_val = (rec.get("licence_area") or "").strip().upper() if rec.get("licence_area") else ""
        if la_val in ("SPD",):
            country = "Scotland"
        elif la_val in ("SPM",):
            country = "Wales"
        else:
            country = fallback_country
            if not la_val:
                rec["licence_area"] = fallback_area

        proj = _raw_row_to_project(
            rec,
            developer="SPEN",
            default_region=fallback_area,
            default_country=country,
            docket_profile_url="https://www.spenergynetworks.co.uk/pages/"
                               "long_term_development_statement.aspx",
        )
        if proj:
            projects.append(proj)
    log.info("SPEN LTDS: %d projects parsed from %s", len(projects), path.name)
    return projects


# ---------------------------------------------------------------------------
# NGED — stub (they use CIM XML, not Excel)
# ---------------------------------------------------------------------------
def parse_nged_workbook(path: Path | None = None, **_: Any) -> list[SubstationProject]:
    """NGED (National Grid Electricity Distribution) LTDS parser — STUB.

    NGED publishes CIM RDF/XML (IEC 61970 Stage 1.3 EQ profile), NOT
    Excel. The canonical path is `utils/ltds_cim_ingester.py`.
    """
    log.warning("NGED LTDS parser is not yet implemented — use CIM pipeline for NGED")
    return []


# ---------------------------------------------------------------------------
# NPG — Northern Powergrid
# ---------------------------------------------------------------------------
# Confidence: MEDIUM. Two tabs (NPN = Northeast, NPY = Yorkshire) with
# 'Substation Name' / 'Voltage (kV)' / 'Rating (MVA)' / 'Headroom (MVA)' /
# 'Grid Reference'. The grid-reference column sometimes carries a single
# OSGB letter-pair string like 'NZ 275 643' and sometimes an easting +
# northing pair — we parse both forms in _raw_row_to_project via
# _split_grid_ref.
NPG_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "name": (
        "Substation Name", "Site Name", "Name",
    ),
    "voltage_kv": (
        "Voltage (kV)", "Nominal Voltage (kV)", "Voltage",
    ),
    "firm_capacity_mva": (
        "Rating (MVA)", "Firm Rating (MVA)", "Firm Capacity (MVA)",
        "Winter Firm Rating (MVA)",
    ),
    "demand_headroom_mw": (
        "Headroom (MVA)", "Demand Headroom (MW)", "Headroom (MW)",
        "Available Headroom (MW)",
    ),
    "peak_demand_mw": (
        "Forecast Peak Demand 2028 (MW)", "Peak Demand (MW)",
        "Maximum Demand (MW)",
    ),
    "easting": (
        "Easting", "OS Easting",
    ),
    "northing": (
        "Northing", "OS Northing",
    ),
    "grid_ref": (
        "Grid Reference", "OS Grid Reference", "Location",
    ),
    "licence_area": (
        "Licence Area", "Region",
    ),
    "site_type": (
        "Type", "Site Type", "Category",
    ),
    "external_id": (
        "Site Code", "Asset ID", "Substation ID", "ID",
    ),
    "operation_year": (
        "Commissioning Year",
    ),
}

NPG_SHEET_ALIASES: tuple[str, ...] = (
    "NPN", "NPY", "LTDS", "Substation", "Primary", "BSP",
)


def _split_grid_ref(ref: Any) -> tuple[float | None, float | None]:
    """Turn an OSGB grid reference into (easting, northing) in metres.

    Accepts two forms:
      * Full numeric "425600 563200" or "425600,563200"
      * Letter-pair "NZ 275 643" (6-figure) / "SP01234 56789" (10-figure)

    Returns (None, None) if the string can't be parsed. We intentionally
    keep this lightweight — for DNO workbooks that already provide two
    numeric columns the helper is not called.
    """
    if ref is None:
        return None, None
    s = str(ref).strip()
    if not s:
        return None, None
    # Pure numeric pair
    m = re.match(r"^(-?\d+(?:\.\d+)?)\D+(-?\d+(?:\.\d+)?)$", s)
    if m:
        try:
            return float(m.group(1)), float(m.group(2))
        except ValueError:
            return None, None
    # Letter-pair form: "NZ 275 643"
    m = re.match(r"^([A-Za-z]{2})\s*(\d+)\s*(\d+)$", s)
    if m:
        letters = m.group(1).upper()
        east_part = m.group(2)
        north_part = m.group(3)
        # OSGB letter → origin table (100km squares). Full OSGB conversion is
        # non-trivial; we use a pragmatic first-letter+second-letter lookup
        # covering the UK mainland only.
        first = {
            "H": (0, 1000000), "N": (0, 500000), "S": (0, 0),
            "T": (500000, 0), "O": (500000, 500000),
        }.get(letters[0])
        if first is None:
            return None, None
        col = (ord(letters[1]) - ord("A")) % 5
        row = 4 - ((ord(letters[1]) - ord("A")) // 5)
        # Skip 'I' which OSGB does not use.
        if letters[1] >= "I":
            col = (ord(letters[1]) - ord("A") - 1) % 5
            row = 4 - ((ord(letters[1]) - ord("A") - 1) // 5)
        sq_e = first[0] + col * 100000
        sq_n = first[1] + row * 100000
        # Pad the per-square parts to 5 digits each (100m × 100m best resolution).
        digits = max(len(east_part), len(north_part))
        if digits == 0 or digits > 5:
            return None, None
        scale = 10 ** (5 - digits)
        try:
            return float(sq_e + int(east_part) * scale), float(sq_n + int(north_part) * scale)
        except ValueError:
            return None, None
    return None, None


def parse_npg_workbook(path: Path | None = None, **_: Any) -> list[SubstationProject]:
    """NPG (Northern Powergrid) LTDS parser.

    NPG has two licence areas: NPN (Northeast) and NPY (Yorkshire). Both
    ship in one workbook with separate tabs; the parser picks up rows
    from any tab whose name matches an `NPG_SHEET_ALIASES` entry and uses
    the `Licence Area` column (when present) or the tab name to assign
    the region.

    TODO — column-mapping (candidate aliases):
      * Name            : 'Substation Name'
      * Voltage         : 'Voltage (kV)'
      * Firm capacity   : 'Rating (MVA)'
      * Demand headroom : 'Headroom (MVA)' | 'Demand Headroom (MW)'
      * Coordinates     : 'Easting'+'Northing' | 'Grid Reference'
    """
    if path is None:
        log.warning("NPG LTDS: no path supplied — returning empty")
        return []
    path = Path(path)
    if not path.exists():
        log.warning("NPG LTDS: %s not found — returning empty", path)
        return []

    _load_openpyxl()

    raw_rows = _parse_workbook_pandas(
        path,
        sheet_aliases=NPG_SHEET_ALIASES,
        field_aliases=NPG_FIELD_ALIASES,
    )

    projects: list[SubstationProject] = []
    for rec in raw_rows:
        # Derive easting/northing from grid_ref if the workbook only
        # provides that column.
        if rec.get("easting") is None and rec.get("northing") is None and rec.get("grid_ref"):
            e, n = _split_grid_ref(rec.get("grid_ref"))
            if e is not None and n is not None:
                rec["easting"] = e
                rec["northing"] = n

        # Licence area — from column if present, else from sheet name.
        la = (rec.get("licence_area") or "").strip().upper()
        if la in ("NPN", "NPY"):
            licence_area = la
        else:
            sheet = str(rec.get("_sheet", "")).upper()
            if "NPN" in sheet:
                licence_area = "NPN"
            elif "NPY" in sheet:
                licence_area = "NPY"
            else:
                licence_area = "NPG"
        rec["licence_area"] = licence_area

        proj = _raw_row_to_project(
            rec,
            developer="NPG",
            default_region=licence_area,
            default_country="England",
            docket_profile_url="https://www.northernpowergrid.com/long-term-development-statement",
        )
        if proj:
            projects.append(proj)
    log.info("NPG LTDS: %d projects parsed from %s", len(projects), path.name)
    return projects


# ---------------------------------------------------------------------------
# ENWL — Electricity North West
# ---------------------------------------------------------------------------
# Confidence: MEDIUM. Single licence area. ENWL's LTDS workbook typically
# uses 'BSP & Primary Substations' as the headline tab with clean row-1
# headers.
ENWL_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "name": (
        "Substation Name", "Site Name", "Name", "Primary Substation",
    ),
    "voltage_kv": (
        "Nominal Voltage (kV)", "Voltage (kV)", "Voltage",
    ),
    "firm_capacity_mva": (
        "Firm Capacity (MVA)", "Firm Capacity MVA", "Firm Rating (MVA)",
    ),
    "demand_headroom_mw": (
        "Demand Headroom (MW)", "Headroom (MW)", "Spare Capacity (MW)",
        "Available Demand Headroom (MW)",
    ),
    "peak_demand_mw": (
        "Winter Peak Demand 2024 (MVA)", "Winter Peak Demand (MVA)",
        "Peak Demand (MW)", "Maximum Demand (MW)",
    ),
    "easting": (
        "Easting", "OS Easting", "X",
    ),
    "northing": (
        "Northing", "OS Northing", "Y",
    ),
    "site_type": (
        "Type", "Substation Type", "Site Type", "Category",
    ),
    "external_id": (
        "Site Code", "Substation Code", "Asset ID", "ID",
    ),
    "operation_year": (
        "Commissioning Year",
    ),
    "licence_area": (
        "Licence Area", "Region",
    ),
}

ENWL_SHEET_ALIASES: tuple[str, ...] = (
    "BSP & Primary Substations", "BSP", "Primary", "Grid",
    "Substation", "Network", "Loading",
)


def parse_enwl_workbook(path: Path | None = None, **_: Any) -> list[SubstationProject]:
    """ENWL (Electricity North West) LTDS parser.

    TODO — column-mapping (candidate aliases):
      * Name            : 'Substation Name'
      * Voltage         : 'Nominal Voltage (kV)' | 'Voltage (kV)'
      * Firm capacity   : 'Firm Capacity (MVA)'
      * Demand headroom : 'Demand Headroom (MW)' | 'Headroom (MW)'
      * Coordinates     : 'Easting' + 'Northing'
    """
    if path is None:
        log.warning("ENWL LTDS: no path supplied — returning empty")
        return []
    path = Path(path)
    if not path.exists():
        log.warning("ENWL LTDS: %s not found — returning empty", path)
        return []

    _load_openpyxl()

    raw_rows = _parse_workbook_pandas(
        path,
        sheet_aliases=ENWL_SHEET_ALIASES,
        field_aliases=ENWL_FIELD_ALIASES,
    )

    projects: list[SubstationProject] = []
    for rec in raw_rows:
        if not rec.get("licence_area"):
            rec["licence_area"] = "ENWL"
        proj = _raw_row_to_project(
            rec,
            developer="ENWL",
            default_region="ENWL",
            default_country="England",
            docket_profile_url="https://www.enwl.co.uk/globalassets/connections/"
                               "application-guidance/long-term-development-statement-ltds/",
        )
        if proj:
            projects.append(proj)
    log.info("ENWL LTDS: %d projects parsed from %s", len(projects), path.name)
    return projects


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------
DNO_DISPATCHERS: dict[str, Callable[..., list[SubstationProject]]] = {
    "UKPN": parse_ukpn_workbook,
    "SSEN": parse_ssen_workbook,
    "SPEN": parse_spen_workbook,
    "NGED": parse_nged_workbook,
    "NPG":  parse_npg_workbook,
    "ENWL": parse_enwl_workbook,
}


def parse_ltds_workbook(dno: str, path: Path | None = None, **kwargs: Any) -> list[SubstationProject]:
    """Dispatch to the DNO-specific parser.

    Args:
        dno:   One of UKPN, SSEN, SPEN, NGED, NPG, ENWL (case-insensitive).
        path:  Path to the .xlsx workbook (may be None for online-only DNOs).
        **kwargs: Passed through to the dispatcher — e.g. `records=` for UKPN.

    Raises:
        ValueError if `dno` is not a known licensee.
    """
    key = dno.upper().strip()
    if key not in DNO_DISPATCHERS:
        raise ValueError(f"Unknown DNO '{dno}'; expected one of {list(DNO_DISPATCHERS)}")
    return DNO_DISPATCHERS[key](path=path, **kwargs)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _safe_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _extract_latlon(rec: dict) -> tuple[float | None, float | None]:
    """Try the common UKPN ODS coordinate carriers in priority order."""
    geo = rec.get("spatial_coordinates") or rec.get("geo_point_2d")
    if isinstance(geo, dict):
        lat = geo.get("lat")
        lon = geo.get("lon") or geo.get("lng")
        if lat and lon:
            return float(lat), float(lon)
    shape = rec.get("geo_shape")
    if isinstance(shape, dict):
        coords = (shape.get("geometry") or {}).get("coordinates")
        if isinstance(coords, list) and len(coords) >= 2:
            return float(coords[1]), float(coords[0])
    return None, None


def _canonical_id(developer: str, name: str, year: int | None) -> str:
    """Build a stable SUB-<dev>-<slug>-<yr> project id."""
    slug = re.sub(r"[^A-Z0-9]+", "-", name.upper()).strip("-")
    slug = slug[:32] if slug else "UNKNOWN"
    yr = str(year) if year else "LIVE"
    return f"SUB-{developer}-{slug}-{yr}"


def _find_header_row(sheet, candidates: tuple[str, ...], max_scan: int = 10) -> tuple[int, list[str]]:
    """Scan the top `max_scan` rows for one that contains any of `candidates`.

    Returns (row_index, normalised_headers_lower). Returns (0, []) if not found.
    """
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if any(any(cand in c for cand in candidates) for c in cells):
            return row_idx, cells
    return 0, []
