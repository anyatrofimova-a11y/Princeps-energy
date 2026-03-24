"""
export_engine.py — Comprehensive export engine for Princeps site assessments.

Three output formats:
  1. generate_site_report_html()  — Branded HTML report (printable A4 PDF-ready)
  2. generate_excel_export()      — Multi-sheet Excel workbook (openpyxl)
  3. generate_csv_export()        — Flat CSV for quick data interchange

Designed for consultancy-grade deliverables: branded, structured, print-ready.
"""

from __future__ import annotations

import csv
import io
import logging
import math
from datetime import datetime, timezone
from typing import Any

log = logging.getLogger("princeps.export_engine")

# ---------------------------------------------------------------------------
# Brand constants
# ---------------------------------------------------------------------------
GOLD = "#F5B731"
GOLD_LIGHT = "#FBD978"
GOLD_MUTED = "#F5B73133"
INK = "#0F1318"
INK_LIGHT = "#3A3F47"
BG = "#F2F3F5"
BG_ALT = "#E8E9EC"
WHITE = "#FFFFFF"
GREEN = "#22C55E"
GREEN_BG = "#DCFCE7"
AMBER = "#F59E0B"
AMBER_BG = "#FEF3C7"
RED = "#EF4444"
RED_BG = "#FEE2E2"
GREY = "#6B7280"
GREY_LIGHT = "#9CA3AF"
BORDER = "#D1D5DB"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe(data: dict, *keys, default="—"):
    """Safely traverse nested dict keys."""
    v = data
    for k in keys:
        if isinstance(v, dict):
            v = v.get(k, None)
        else:
            return default
    return v if v is not None else default


def _fmt_money(v, prefix="£") -> str:
    if v is None or v == "—":
        return "—"
    try:
        n = float(v)
        if abs(n) >= 1_000_000:
            return f"{prefix}{n / 1_000_000:,.1f}M"
        if abs(n) >= 1_000:
            return f"{prefix}{n / 1_000:,.0f}k"
        return f"{prefix}{n:,.0f}"
    except (ValueError, TypeError):
        return str(v)


def _fmt_pct(v) -> str:
    if v is None or v == "—":
        return "—"
    try:
        return f"{float(v):.1f}%"
    except (ValueError, TypeError):
        return str(v)


def _score_colour(score) -> str:
    """Return hex colour for a 0-100 score."""
    try:
        s = float(score)
    except (ValueError, TypeError):
        return GREY
    if s >= 70:
        return GREEN
    if s >= 40:
        return AMBER
    return RED


def _verdict_colour(verdict: str) -> str:
    v = str(verdict).upper().replace("-", "").replace("_", "").replace(" ", "")
    if v in ("GO", "PROCEED"):
        return GREEN
    if v in ("CAUTION", "CONDITIONAL"):
        return AMBER
    return RED


def _risk_score(likelihood, impact) -> int:
    try:
        return int(likelihood) * int(impact)
    except (ValueError, TypeError):
        return 0


def _risk_colour(score: int) -> str:
    if score >= 15:
        return RED
    if score >= 8:
        return AMBER
    return GREEN


# ---------------------------------------------------------------------------
# 1. HTML Report
# ---------------------------------------------------------------------------

def generate_site_report_html(site_data: dict) -> str:
    """
    Generate a branded HTML report for a site assessment.

    site_data keys (all optional, graceful fallback to '—'):
        name, lat, lon, area_ha, proposed_use, capacity_mw, verdict, score,
        recommendation, dno_area,
        grid: {nearest_substations: [...], headroom_mw, queue_position,
               cost_p10, cost_p50, cost_p90, timeline_months, voltage_kv},
        environmental: {constraints: [...], distances: {...},
                        required_surveys: [...], estimated_cost},
        financial: {capex, opex_annual, revenue_annual, irr, payback_years,
                    npv, lcoe, capex_breakdown: {equipment, civils, grid, ...}},
        planning: {lpa, approval_rate, eia_required, bng_obligation,
                   nsip_threshold, planning_risk},
        risks: [{risk, likelihood, impact, category, mitigation}, ...],
        next_steps: [str, ...],
        data_sources: [str, ...],
    """
    d = site_data or {}
    now = datetime.now(timezone.utc).strftime("%d %B %Y")
    name = _safe(d, "name", default="Unnamed Site")
    verdict = _safe(d, "verdict", default="PENDING")
    score = _safe(d, "score", default=0)
    recommendation = _safe(d, "recommendation", default="Further assessment required.")

    # Grid section
    g = d.get("grid") or {}
    substations = g.get("nearest_substations") or []
    # Environmental
    e = d.get("environmental") or {}
    constraints = e.get("constraints") or []
    surveys = e.get("required_surveys") or []
    # Financial
    f = d.get("financial") or {}
    capex_breakdown = f.get("capex_breakdown") or {}
    # Planning
    p = d.get("planning") or {}
    # Risks
    risks = d.get("risks") or _default_risks()
    # Next steps
    steps = d.get("next_steps") or _default_next_steps()
    # Data sources
    sources = d.get("data_sources") or _default_data_sources()

    # Build inline SVG charts
    score_gauge_svg = _build_score_gauge(score)
    capex_bar_svg = _build_capex_bar_chart(capex_breakdown) if capex_breakdown else ""
    risk_matrix_svg = _build_risk_matrix_svg(risks)

    # Substation table rows
    sub_rows = ""
    for s in substations[:6]:
        if isinstance(s, dict):
            sub_rows += f"""<tr>
                <td>{s.get('name', '—')}</td>
                <td>{s.get('voltage_kv', '—')} kV</td>
                <td>{s.get('distance_km', '—')} km</td>
                <td>{s.get('headroom_mw', '—')} MW</td>
                <td>{s.get('queue', '—')}</td>
            </tr>"""

    # Constraint rows
    constraint_rows = ""
    for c in constraints[:10]:
        if isinstance(c, dict):
            constraint_rows += f"""<tr>
                <td>{c.get('type', '—')}</td>
                <td>{c.get('name', '—')}</td>
                <td>{c.get('distance_m', '—')} m</td>
                <td><span class="badge {'badge-red' if c.get('blocking') else 'badge-amber'}">{
                    'Blocking' if c.get('blocking') else 'Advisory'}</span></td>
            </tr>"""
        elif isinstance(c, str):
            constraint_rows += f"""<tr><td colspan="4">{c}</td></tr>"""

    # Risk rows
    risk_rows = ""
    for r in risks[:10]:
        rs = _risk_score(r.get("likelihood", 0), r.get("impact", 0))
        rc = _risk_colour(rs)
        risk_rows += f"""<tr>
            <td>{r.get('risk', '—')}</td>
            <td>{r.get('category', '—')}</td>
            <td class="center">{r.get('likelihood', '—')}</td>
            <td class="center">{r.get('impact', '—')}</td>
            <td class="center" style="background:{rc}22;color:{rc};font-weight:700;">{rs}</td>
            <td>{r.get('mitigation', '—')}</td>
        </tr>"""

    # Survey rows
    survey_rows = ""
    for sv in surveys[:8]:
        if isinstance(sv, dict):
            survey_rows += f"""<tr>
                <td>{sv.get('survey', '—')}</td>
                <td>{_fmt_money(sv.get('cost'))}</td>
                <td>{sv.get('timeline', '—')}</td>
            </tr>"""
        elif isinstance(sv, str):
            survey_rows += f"""<tr><td colspan="3">{sv}</td></tr>"""

    # Next steps
    steps_html = ""
    for i, st in enumerate(steps[:10], 1):
        steps_html += f"""<div class="step-item">
            <div class="step-number">{i}</div>
            <div class="step-text">{st}</div>
        </div>"""

    # Data sources
    sources_html = "".join(f"<li>{s}</li>" for s in sources[:15])

    # Cost P10/P50/P90 bar
    cost_range_svg = _build_cost_range_svg(
        _safe(g, "cost_p10"), _safe(g, "cost_p50"), _safe(g, "cost_p90")
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Princeps — Site Assessment Report — {_esc(name)}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;1,9..40,400&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">
<style>
/* ── Reset & Base ────────────────────────────── */
*, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
html {{ font-size: 14px; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
body {{
    font-family: 'DM Sans', -apple-system, BlinkMacSystemFont, sans-serif;
    color: {INK};
    background: {WHITE};
    line-height: 1.6;
}}

/* ── Print / A4 Layout ───────────────────────── */
@page {{
    size: A4;
    margin: 18mm 15mm 20mm 15mm;
}}
@media print {{
    .page-break {{ page-break-before: always; }}
    .no-print {{ display: none; }}
    body {{ font-size: 12px; }}
}}
.report {{
    max-width: 210mm;
    margin: 0 auto;
    padding: 40px 36px;
}}

/* ── Cover / Header ──────────────────────────── */
.cover {{
    border-bottom: 3px solid {GOLD};
    padding-bottom: 32px;
    margin-bottom: 36px;
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
}}
.cover-left {{ flex: 1; }}
.logo-mark {{
    width: 48px; height: 48px;
    background: {GOLD};
    border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
    font-family: 'DM Sans', sans-serif;
    font-weight: 700; font-size: 28px; color: {INK};
    margin-bottom: 16px;
}}
.report-title {{
    font-size: 28px; font-weight: 700; color: {INK};
    margin-bottom: 4px; letter-spacing: -0.5px;
}}
.report-subtitle {{
    font-size: 15px; color: {GREY}; font-weight: 400;
}}
.cover-right {{
    text-align: right; font-size: 13px; color: {GREY};
    line-height: 1.8;
}}
.cover-right strong {{ color: {INK}; }}

/* ── Section headings ────────────────────────── */
.section {{
    margin-bottom: 32px;
}}
.section-heading {{
    font-size: 18px; font-weight: 700; color: {INK};
    border-left: 4px solid {GOLD};
    padding-left: 12px;
    margin-bottom: 16px;
}}
.section-number {{
    color: {GOLD}; margin-right: 6px; font-weight: 600;
}}

/* ── Executive Summary ───────────────────────── */
.exec-grid {{
    display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px;
    margin-bottom: 20px;
}}
.exec-card {{
    background: {BG};
    border-radius: 10px;
    padding: 20px;
    text-align: center;
}}
.exec-card.verdict {{
    border: 2px solid {_verdict_colour(verdict)};
}}
.exec-label {{
    font-size: 11px; text-transform: uppercase; letter-spacing: 1px;
    color: {GREY}; margin-bottom: 6px; font-weight: 500;
}}
.exec-value {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 26px; font-weight: 700;
}}
.verdict-text {{
    color: {_verdict_colour(verdict)};
}}
.recommendation-box {{
    background: {GOLD}18;
    border-left: 4px solid {GOLD};
    padding: 14px 18px;
    border-radius: 0 8px 8px 0;
    font-size: 14px;
    margin-top: 12px;
}}

/* ── Tables ───────────────────────────────────── */
table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
    margin-bottom: 16px;
}}
th {{
    background: {INK};
    color: {WHITE};
    font-weight: 600;
    text-align: left;
    padding: 10px 12px;
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}}
td {{
    padding: 9px 12px;
    border-bottom: 1px solid {BORDER};
    font-family: 'JetBrains Mono', monospace;
    font-size: 12px;
}}
tr:nth-child(even) td {{
    background: {BG};
}}
tr:hover td {{
    background: {GOLD}0D;
}}
.center {{ text-align: center; }}

/* ── Key-value pairs ──────────────────────────── */
.kv-grid {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 8px 32px;
    margin-bottom: 16px;
}}
.kv {{
    display: flex;
    justify-content: space-between;
    padding: 7px 0;
    border-bottom: 1px solid {BORDER};
}}
.kv-label {{
    color: {GREY}; font-size: 13px; font-weight: 500;
}}
.kv-value {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 13px; font-weight: 600; color: {INK};
}}

/* ── Badges ───────────────────────────────────── */
.badge {{
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}}
.badge-green {{ background: {GREEN_BG}; color: {GREEN}; }}
.badge-amber {{ background: {AMBER_BG}; color: {AMBER}; }}
.badge-red {{ background: {RED_BG}; color: {RED}; }}

/* ── Steps ────────────────────────────────────── */
.step-item {{
    display: flex; align-items: flex-start; gap: 12px;
    margin-bottom: 10px;
}}
.step-number {{
    flex-shrink: 0;
    width: 28px; height: 28px;
    background: {GOLD};
    color: {INK};
    font-weight: 700; font-size: 14px;
    border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
}}
.step-text {{
    font-size: 13px; line-height: 1.5; padding-top: 4px;
}}

/* ── Charts container ─────────────────────────── */
.chart-container {{
    display: flex;
    justify-content: center;
    margin: 16px 0;
}}
.chart-side {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 24px;
    align-items: start;
}}

/* ── Footer ───────────────────────────────────── */
.footer {{
    margin-top: 48px;
    padding-top: 16px;
    border-top: 2px solid {GOLD};
    display: flex;
    justify-content: space-between;
    font-size: 11px;
    color: {GREY};
}}
.footer-logo {{
    display: inline-flex; align-items: center; gap: 6px;
    font-weight: 700; color: {INK}; font-size: 12px;
}}
.footer-logo-mark {{
    width: 18px; height: 18px; background: {GOLD};
    border-radius: 4px; display: flex; align-items: center;
    justify-content: center; font-size: 11px; font-weight: 700; color: {INK};
}}

/* ── Appendix ─────────────────────────────────── */
.appendix-list {{
    list-style: none; padding: 0;
}}
.appendix-list li {{
    padding: 6px 0;
    border-bottom: 1px solid {BORDER};
    font-size: 12px;
    color: {INK_LIGHT};
}}
.appendix-list li::before {{
    content: "→ ";
    color: {GOLD};
    font-weight: 700;
}}
</style>
</head>
<body>
<div class="report">

    <!-- ═══ COVER ═══ -->
    <div class="cover">
        <div class="cover-left">
            <div class="logo-mark">P</div>
            <div class="report-title">Site Assessment Report</div>
            <div class="report-subtitle">{_esc(name)}</div>
        </div>
        <div class="cover-right">
            <strong>PRINCEPS</strong><br>
            Energy Infrastructure Intelligence<br><br>
            Report generated: {now}<br>
            Classification: <strong>CONFIDENTIAL</strong><br>
            Reference: PR-{abs(hash(name)) % 100000:05d}
        </div>
    </div>

    <!-- ═══ 1. EXECUTIVE SUMMARY ═══ -->
    <div class="section">
        <h2 class="section-heading"><span class="section-number">01</span> Executive Summary</h2>
        <div class="exec-grid">
            <div class="exec-card verdict">
                <div class="exec-label">Verdict</div>
                <div class="exec-value verdict-text">{_esc(str(verdict).upper())}</div>
            </div>
            <div class="exec-card">
                <div class="exec-label">Overall Score</div>
                <div class="exec-value" style="color:{_score_colour(score)}">{score}<span style="font-size:14px;color:{GREY};">/100</span></div>
            </div>
            <div class="exec-card">
                <div class="exec-label">Proposed Capacity</div>
                <div class="exec-value">{_safe(d, 'capacity_mw', default='—')}<span style="font-size:14px;color:{GREY};"> MW</span></div>
            </div>
        </div>
        <div class="chart-container">
            {score_gauge_svg}
        </div>
        <div class="kv-grid">
            <div class="kv"><span class="kv-label">Site Name</span><span class="kv-value">{_esc(name)}</span></div>
            <div class="kv"><span class="kv-label">Proposed Use</span><span class="kv-value">{_esc(str(_safe(d, 'proposed_use')))}</span></div>
            <div class="kv"><span class="kv-label">Grid Headroom</span><span class="kv-value">{_safe(g, 'headroom_mw')} MW</span></div>
            <div class="kv"><span class="kv-label">Est. Connection Cost</span><span class="kv-value">{_fmt_money(_safe(g, 'cost_p50'))}</span></div>
            <div class="kv"><span class="kv-label">Est. IRR</span><span class="kv-value">{_fmt_pct(_safe(f, 'irr'))}</span></div>
            <div class="kv"><span class="kv-label">Payback Period</span><span class="kv-value">{_safe(f, 'payback_years')} yrs</span></div>
        </div>
        <div class="recommendation-box">
            <strong>Recommendation:</strong> {_esc(str(recommendation))}
        </div>
    </div>

    <!-- ═══ 2. SITE OVERVIEW ═══ -->
    <div class="section">
        <h2 class="section-heading"><span class="section-number">02</span> Site Overview</h2>
        <div class="kv-grid">
            <div class="kv"><span class="kv-label">Latitude</span><span class="kv-value">{_safe(d, 'lat')}</span></div>
            <div class="kv"><span class="kv-label">Longitude</span><span class="kv-value">{_safe(d, 'lon')}</span></div>
            <div class="kv"><span class="kv-label">Site Area</span><span class="kv-value">{_safe(d, 'area_ha')} ha</span></div>
            <div class="kv"><span class="kv-label">DNO Area</span><span class="kv-value">{_esc(str(_safe(d, 'dno_area')))}</span></div>
            <div class="kv"><span class="kv-label">Proposed Use</span><span class="kv-value">{_esc(str(_safe(d, 'proposed_use')))}</span></div>
            <div class="kv"><span class="kv-label">Capacity</span><span class="kv-value">{_safe(d, 'capacity_mw')} MW</span></div>
            <div class="kv"><span class="kv-label">Land Classification</span><span class="kv-value">{_esc(str(_safe(d, 'land_class', default='—')))}</span></div>
            <div class="kv"><span class="kv-label">Elevation</span><span class="kv-value">{_safe(d, 'elevation_m', default='—')} m AOD</span></div>
        </div>
    </div>

    <!-- ═══ 3. GRID CONNECTION ═══ -->
    <div class="section page-break">
        <h2 class="section-heading"><span class="section-number">03</span> Grid Connection Assessment</h2>
        <div class="kv-grid">
            <div class="kv"><span class="kv-label">Available Headroom</span><span class="kv-value">{_safe(g, 'headroom_mw')} MW</span></div>
            <div class="kv"><span class="kv-label">Connection Voltage</span><span class="kv-value">{_safe(g, 'voltage_kv')} kV</span></div>
            <div class="kv"><span class="kv-label">Queue Position</span><span class="kv-value">{_safe(g, 'queue_position')}</span></div>
            <div class="kv"><span class="kv-label">Est. Timeline</span><span class="kv-value">{_safe(g, 'timeline_months')} months</span></div>
        </div>
        {"<h3 style='font-size:14px;margin:16px 0 8px;'>Nearest Substations</h3>" if sub_rows else ""}
        {"<table><tr><th>Substation</th><th>Voltage</th><th>Distance</th><th>Headroom</th><th>Queue</th></tr>" + sub_rows + "</table>" if sub_rows else ""}
        <h3 style="font-size:14px;margin:16px 0 8px;">Connection Cost Estimate (P10 / P50 / P90)</h3>
        <div class="kv-grid">
            <div class="kv"><span class="kv-label">P10 (Optimistic)</span><span class="kv-value">{_fmt_money(_safe(g, 'cost_p10'))}</span></div>
            <div class="kv"><span class="kv-label">P50 (Central)</span><span class="kv-value" style="font-size:15px;">{_fmt_money(_safe(g, 'cost_p50'))}</span></div>
            <div class="kv"><span class="kv-label">P90 (Conservative)</span><span class="kv-value">{_fmt_money(_safe(g, 'cost_p90'))}</span></div>
            <div class="kv"><span class="kv-label"></span><span class="kv-value"></span></div>
        </div>
        {f'<div class="chart-container">{cost_range_svg}</div>' if cost_range_svg else ""}
    </div>

    <!-- ═══ 4. ENVIRONMENTAL ═══ -->
    <div class="section">
        <h2 class="section-heading"><span class="section-number">04</span> Environmental Assessment</h2>
        {"<table><tr><th>Constraint Type</th><th>Name</th><th>Distance</th><th>Status</th></tr>" + constraint_rows + "</table>" if constraint_rows else '<p style="color:' + GREY + ';">No environmental constraints identified within search radius.</p>'}
        {"<h3 style='font-size:14px;margin:16px 0 8px;'>Required Surveys</h3>" if survey_rows else ""}
        {"<table><tr><th>Survey</th><th>Est. Cost</th><th>Timeline</th></tr>" + survey_rows + "</table>" if survey_rows else ""}
        <div class="kv-grid">
            <div class="kv"><span class="kv-label">Total Environmental Survey Cost</span><span class="kv-value">{_fmt_money(_safe(e, 'estimated_cost'))}</span></div>
            <div class="kv"><span class="kv-label">Flood Risk Zone</span><span class="kv-value">{_esc(str(_safe(e, 'flood_zone', default='—')))}</span></div>
        </div>
    </div>

    <!-- ═══ 5. FINANCIAL SUMMARY ═══ -->
    <div class="section page-break">
        <h2 class="section-heading"><span class="section-number">05</span> Financial Summary</h2>
        <div class="exec-grid">
            <div class="exec-card">
                <div class="exec-label">Total CapEx</div>
                <div class="exec-value">{_fmt_money(_safe(f, 'capex'))}</div>
            </div>
            <div class="exec-card">
                <div class="exec-label">Annual OpEx</div>
                <div class="exec-value">{_fmt_money(_safe(f, 'opex_annual'))}</div>
            </div>
            <div class="exec-card">
                <div class="exec-label">Annual Revenue</div>
                <div class="exec-value">{_fmt_money(_safe(f, 'revenue_annual'))}</div>
            </div>
        </div>
        <div class="kv-grid">
            <div class="kv"><span class="kv-label">Internal Rate of Return</span><span class="kv-value" style="color:{_score_colour(float(_safe(f, 'irr', default=0)) if _safe(f, 'irr', default=0) != '—' else 0)}">{_fmt_pct(_safe(f, 'irr'))}</span></div>
            <div class="kv"><span class="kv-label">Net Present Value</span><span class="kv-value">{_fmt_money(_safe(f, 'npv'))}</span></div>
            <div class="kv"><span class="kv-label">Payback Period</span><span class="kv-value">{_safe(f, 'payback_years')} years</span></div>
            <div class="kv"><span class="kv-label">LCOE</span><span class="kv-value">{_safe(f, 'lcoe')} £/MWh</span></div>
        </div>
        {f'<h3 style="font-size:14px;margin:16px 0 8px;">CapEx Breakdown</h3><div class="chart-container">{capex_bar_svg}</div>' if capex_bar_svg else ""}
    </div>

    <!-- ═══ 6. PLANNING ASSESSMENT ═══ -->
    <div class="section">
        <h2 class="section-heading"><span class="section-number">06</span> Planning Assessment</h2>
        <div class="kv-grid">
            <div class="kv"><span class="kv-label">Local Planning Authority</span><span class="kv-value">{_esc(str(_safe(p, 'lpa')))}</span></div>
            <div class="kv"><span class="kv-label">LPA Approval Rate</span><span class="kv-value">{_fmt_pct(_safe(p, 'approval_rate'))}</span></div>
            <div class="kv"><span class="kv-label">EIA Required</span><span class="kv-value">{_yes_no(_safe(p, 'eia_required'))}</span></div>
            <div class="kv"><span class="kv-label">BNG Obligation</span><span class="kv-value">{_esc(str(_safe(p, 'bng_obligation')))}</span></div>
            <div class="kv"><span class="kv-label">NSIP Threshold</span><span class="kv-value">{_esc(str(_safe(p, 'nsip_threshold')))}</span></div>
            <div class="kv"><span class="kv-label">Planning Risk</span><span class="kv-value">{_esc(str(_safe(p, 'planning_risk')))}</span></div>
        </div>
    </div>

    <!-- ═══ 7. RISK REGISTER ═══ -->
    <div class="section page-break">
        <h2 class="section-heading"><span class="section-number">07</span> Risk Register</h2>
        <p style="font-size:12px;color:{GREY};margin-bottom:12px;">
            Likelihood (1–5) &times; Impact (1–5). Score &ge;15 = Critical, 8–14 = Moderate, &lt;8 = Low.
        </p>
        <table>
            <tr><th>Risk</th><th>Category</th><th>L</th><th>I</th><th>Score</th><th>Mitigation</th></tr>
            {risk_rows}
        </table>
        {f'<div class="chart-container">{risk_matrix_svg}</div>' if risk_matrix_svg else ""}
    </div>

    <!-- ═══ 8. NEXT STEPS ═══ -->
    <div class="section">
        <h2 class="section-heading"><span class="section-number">08</span> Recommended Next Steps</h2>
        {steps_html}
    </div>

    <!-- ═══ 9. APPENDIX ═══ -->
    <div class="section page-break">
        <h2 class="section-heading"><span class="section-number">09</span> Appendix: Data Sources &amp; Methodology</h2>
        <ul class="appendix-list">
            {sources_html}
        </ul>
        <p style="font-size:11px;color:{GREY};margin-top:16px;">
            This report was generated by Princeps automated site assessment engine.
            All data is indicative and should be validated through detailed surveys
            and professional consultancy before investment decisions are made.
            Grid connection costs are based on published DNO rate cards and may vary
            following formal application.
        </p>
    </div>

    <!-- ═══ FOOTER ═══ -->
    <div class="footer">
        <div>
            <span class="footer-logo">
                <span class="footer-logo-mark">P</span> PRINCEPS
            </span>
            &nbsp;&middot;&nbsp; Energy Infrastructure Intelligence
        </div>
        <div>Confidential &middot; {now} &middot; Page <span class="page-num"></span></div>
    </div>

</div>
</body>
</html>"""

    return html


# ---------------------------------------------------------------------------
# HTML helpers
# ---------------------------------------------------------------------------

def _esc(text: str) -> str:
    """Basic HTML escaping."""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _yes_no(val) -> str:
    if val is None or val == "—":
        return "—"
    if isinstance(val, bool):
        return "Yes" if val else "No"
    s = str(val).lower()
    if s in ("true", "yes", "1"):
        return "Yes"
    if s in ("false", "no", "0"):
        return "No"
    return str(val)


# ---------------------------------------------------------------------------
# Inline SVG chart builders
# ---------------------------------------------------------------------------

def _build_score_gauge(score) -> str:
    """Semi-circular gauge showing 0-100 score."""
    try:
        s = float(score)
    except (ValueError, TypeError):
        s = 0
    s = max(0, min(100, s))

    # Arc calculation: 180-degree arc from left to right
    # Sweep from 180° to 0° (left to right across the top)
    angle = math.pi * (1 - s / 100)
    cx, cy, r = 120, 110, 90
    # Background arc (full semicircle)
    bg_x1, bg_y1 = cx - r, cy
    bg_x2, bg_y2 = cx + r, cy
    # Score arc endpoint
    ex = cx + r * math.cos(angle)
    ey = cy - r * math.sin(angle)
    large_arc = 1 if s > 50 else 0
    colour = _score_colour(s)

    return f"""<svg width="240" height="140" viewBox="0 0 240 140" xmlns="http://www.w3.org/2000/svg">
        <path d="M {bg_x1},{bg_y1} A {r},{r} 0 1,1 {bg_x2},{bg_y2}"
              fill="none" stroke="{BG_ALT}" stroke-width="16" stroke-linecap="round"/>
        <path d="M {bg_x1},{bg_y1} A {r},{r} 0 {large_arc},1 {ex:.1f},{ey:.1f}"
              fill="none" stroke="{colour}" stroke-width="16" stroke-linecap="round"/>
        <text x="{cx}" y="{cy + 5}" text-anchor="middle"
              font-family="JetBrains Mono, monospace" font-size="32" font-weight="700"
              fill="{INK}">{s:.0f}</text>
        <text x="{cx}" y="{cy + 22}" text-anchor="middle"
              font-family="DM Sans, sans-serif" font-size="11" fill="{GREY}">out of 100</text>
    </svg>"""


def _build_capex_bar_chart(breakdown: dict) -> str:
    """Horizontal bar chart for CapEx breakdown."""
    if not breakdown:
        return ""
    items = sorted(breakdown.items(), key=lambda x: -(x[1] or 0))[:8]
    if not items:
        return ""
    max_val = max(v for _, v in items if v) or 1

    bar_h = 28
    gap = 6
    label_w = 100
    bar_w = 280
    svg_h = (bar_h + gap) * len(items) + 10
    svg_w = label_w + bar_w + 80

    bars = ""
    for i, (label, val) in enumerate(items):
        y = i * (bar_h + gap) + 5
        w = (val / max_val) * bar_w if val else 0
        bars += f"""
        <text x="{label_w - 8}" y="{y + bar_h / 2 + 4}" text-anchor="end"
              font-family="DM Sans, sans-serif" font-size="11" fill="{INK_LIGHT}">{_esc(str(label).replace('_', ' ').title())}</text>
        <rect x="{label_w}" y="{y}" width="{w:.0f}" height="{bar_h}"
              rx="4" fill="{GOLD}" opacity="0.85"/>
        <text x="{label_w + w + 6:.0f}" y="{y + bar_h / 2 + 4}"
              font-family="JetBrains Mono, monospace" font-size="11" fill="{INK}">{_fmt_money(val)}</text>"""

    return f"""<svg width="{svg_w}" height="{svg_h}" viewBox="0 0 {svg_w} {svg_h}" xmlns="http://www.w3.org/2000/svg">
        {bars}
    </svg>"""


def _build_cost_range_svg(p10, p50, p90) -> str:
    """Horizontal range bar showing P10/P50/P90 cost estimates."""
    vals = []
    for label, v in [("P10", p10), ("P50", p50), ("P90", p90)]:
        try:
            vals.append((label, float(v)))
        except (ValueError, TypeError):
            pass
    if len(vals) < 2:
        return ""

    min_v = vals[0][1]
    max_v = vals[-1][1]
    spread = max_v - min_v if max_v != min_v else 1

    w, h = 420, 70
    pad = 40
    bar_w = w - 2 * pad
    bar_y = 28
    bar_h = 14

    markers = ""
    for label, v in vals:
        x = pad + ((v - min_v) / spread) * bar_w
        colour = GOLD if label == "P50" else GREY_LIGHT
        size = 10 if label == "P50" else 7
        markers += f"""
        <circle cx="{x:.0f}" cy="{bar_y + bar_h / 2}" r="{size}" fill="{colour}" stroke="{INK}" stroke-width="2"/>
        <text x="{x:.0f}" y="{bar_y - 8}" text-anchor="middle"
              font-family="JetBrains Mono, monospace" font-size="10" font-weight="600" fill="{INK}">{label}</text>
        <text x="{x:.0f}" y="{bar_y + bar_h + 16}" text-anchor="middle"
              font-family="JetBrains Mono, monospace" font-size="10" fill="{GREY}">{_fmt_money(v)}</text>"""

    return f"""<svg width="{w}" height="{h}" viewBox="0 0 {w} {h}" xmlns="http://www.w3.org/2000/svg">
        <rect x="{pad}" y="{bar_y}" width="{bar_w}" height="{bar_h}" rx="7" fill="{BG_ALT}"/>
        <rect x="{pad}" y="{bar_y}" width="{bar_w}" height="{bar_h}" rx="7" fill="{GOLD}" opacity="0.2"/>
        {markers}
    </svg>"""


def _build_risk_matrix_svg(risks: list) -> str:
    """5x5 risk matrix (likelihood x impact) with plotted risks."""
    cell = 44
    pad = 50
    w = 5 * cell + pad + 30
    h = 5 * cell + pad + 30

    # Background cells (green/amber/red zones)
    cells = ""
    for li in range(5):
        for im in range(5):
            score = (li + 1) * (im + 1)
            if score >= 15:
                fill = RED_BG
            elif score >= 8:
                fill = AMBER_BG
            else:
                fill = GREEN_BG
            x = pad + im * cell
            y = pad + (4 - li) * cell  # invert y so L=1 is bottom
            cells += f'<rect x="{x}" y="{y}" width="{cell}" height="{cell}" fill="{fill}" stroke="{BORDER}" stroke-width="1"/>'

    # Axis labels
    labels = ""
    for i in range(5):
        # Impact (x-axis)
        labels += f'<text x="{pad + i * cell + cell / 2}" y="{pad + 5 * cell + 16}" text-anchor="middle" font-family="JetBrains Mono, monospace" font-size="10" fill="{GREY}">{i + 1}</text>'
        # Likelihood (y-axis)
        labels += f'<text x="{pad - 8}" y="{pad + (4 - i) * cell + cell / 2 + 4}" text-anchor="end" font-family="JetBrains Mono, monospace" font-size="10" fill="{GREY}">{i + 1}</text>'

    labels += f'<text x="{pad + 5 * cell / 2}" y="{pad + 5 * cell + 30}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="10" fill="{GREY}">Impact</text>'
    labels += f'<text x="12" y="{pad + 5 * cell / 2}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="10" fill="{GREY}" transform="rotate(-90, 12, {pad + 5 * cell / 2})">Likelihood</text>'

    # Plot risks
    dots = ""
    for idx, r in enumerate(risks[:10]):
        try:
            li = int(r.get("likelihood", 0))
            im = int(r.get("impact", 0))
        except (ValueError, TypeError):
            continue
        if not (1 <= li <= 5 and 1 <= im <= 5):
            continue
        x = pad + (im - 1) * cell + cell / 2 + (idx % 3 - 1) * 8
        y = pad + (5 - li) * cell + cell / 2 + (idx // 3 - 1) * 8
        sc = li * im
        col = RED if sc >= 15 else (AMBER if sc >= 8 else GREEN)
        dots += f'<circle cx="{x:.0f}" cy="{y:.0f}" r="6" fill="{col}" stroke="{WHITE}" stroke-width="1.5"/>'
        dots += f'<text x="{x:.0f}" y="{y + 3:.0f}" text-anchor="middle" font-family="JetBrains Mono, monospace" font-size="7" fill="{WHITE}" font-weight="700">{idx + 1}</text>'

    return f"""<svg width="{w}" height="{h}" viewBox="0 0 {w} {h}" xmlns="http://www.w3.org/2000/svg">
        {cells}{labels}{dots}
    </svg>"""


# ---------------------------------------------------------------------------
# Default data for missing sections
# ---------------------------------------------------------------------------

def _default_risks() -> list:
    return [
        {"risk": "Grid capacity insufficient for proposed scheme",
         "likelihood": 2, "impact": 5, "category": "Grid",
         "mitigation": "Explore flexible connection or phased energisation"},
        {"risk": "Connection timeline exceeds 36 months",
         "likelihood": 3, "impact": 4, "category": "Grid",
         "mitigation": "Early pre-application engagement with DNO"},
        {"risk": "Planning application refused",
         "likelihood": 2, "impact": 5, "category": "Planning",
         "mitigation": "Pre-app consultation, community engagement strategy"},
        {"risk": "Protected species discovered on site",
         "likelihood": 3, "impact": 3, "category": "Environmental",
         "mitigation": "Commission Phase 1 habitat survey early"},
        {"risk": "CapEx overrun exceeds contingency",
         "likelihood": 3, "impact": 3, "category": "Financial",
         "mitigation": "Fixed-price EPC contract with performance guarantees"},
    ]


def _default_next_steps() -> list:
    return [
        "Commission topographical survey and ground investigation",
        "Submit grid connection pre-application to DNO",
        "Engage with Local Planning Authority for pre-application advice",
        "Commission Phase 1 ecological survey",
        "Obtain detailed flood risk assessment if in Flood Zone 2/3",
        "Prepare outline financial model for investment committee",
        "Secure site option agreement with landowner",
        "Appoint planning consultant for EIA screening opinion",
    ]


def _default_data_sources() -> list:
    return [
        "National Grid ESO Embedded Capacity Register (ECR)",
        "DNO Long Term Development Statements (LTDS)",
        "Ofgem Renewable Energy Planning Database (REPD)",
        "Environment Agency Flood Risk Mapping",
        "Natural England MAGIC Constraints Data",
        "Ordnance Survey Open Data (boundaries, terrain)",
        "BMRS / Elexon settlement data",
        "Google Earth Engine (DynamicWorld, ERA5, NASADEM, Sentinel-2)",
        "OpenStreetMap power infrastructure",
        "Land Registry INSPIRE Index Polygons",
        "NESO Future Energy Scenarios 2024",
        "Princeps grid connection cost model (based on published DNO rate cards)",
        "Princeps environmental constraint engine (500m–2km search radii)",
    ]


# ---------------------------------------------------------------------------
# 2. Excel Export
# ---------------------------------------------------------------------------

def generate_excel_export(sites: list[dict]) -> bytes:
    """
    Generate a multi-sheet Excel workbook for site comparison.

    Each item in `sites` should have the same shape as generate_site_report_html's
    site_data dict.

    Returns bytes of the .xlsx file.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── Style definitions ──────────────────────────────────
    gold_fill = PatternFill(start_color="F5B731", end_color="F5B731", fill_type="solid")
    ink_fill = PatternFill(start_color="0F1318", end_color="0F1318", fill_type="solid")
    bg_fill = PatternFill(start_color="F2F3F5", end_color="F2F3F5", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    amber_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    title_font = Font(name="DM Sans", bold=True, size=14, color="0F1318")
    header_font = Font(name="DM Sans", bold=True, size=10, color="FFFFFF")
    label_font = Font(name="DM Sans", bold=True, size=10, color="3A3F47")
    value_font = Font(name="JetBrains Mono", size=10, color="0F1318")
    section_font = Font(name="DM Sans", bold=True, size=11, color="F5B731")
    section_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")

    thin_border = Border(bottom=Side(style="thin", color="D1D5DB"))
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")

    def _brand_header(ws, row: int, cols: int):
        """Add Princeps branding header to a sheet."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        cell = ws.cell(row=row, column=1, value="PRINCEPS — Site Assessment Export")
        cell.font = title_font
        cell.fill = gold_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 32
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=cols)
        ts = ws.cell(row=row + 1, column=1,
                     value=f"Generated: {datetime.now(timezone.utc).strftime('%d %B %Y %H:%M UTC')}")
        ts.font = Font(name="DM Sans", size=9, color="6B7280")
        return row + 3  # next usable row

    def _header_row(ws, row: int, labels: list):
        for col, label in enumerate(labels, 1):
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = header_font
            cell.fill = ink_fill
            cell.alignment = header_align
        ws.row_dimensions[row].height = 28

    def _auto_width(ws, min_w=12, max_w=40):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), max_w)

    if not sites:
        sites = [{}]

    site_names = [s.get("name", f"Site {i+1}") for i, s in enumerate(sites)]
    num_sites = len(sites)

    # ═══════════════════════════════════════════════════════
    # Sheet 1: Site Comparison
    # ═══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Site Comparison"

    total_cols = num_sites + 1
    row = _brand_header(ws1, 1, total_cols)

    # Header
    _header_row(ws1, row, ["Metric"] + site_names)
    row += 1

    # Metrics to compare
    comparison_metrics = [
        ("General", [
            ("Verdict", lambda s: _safe(s, "verdict")),
            ("Overall Score", lambda s: _safe(s, "score")),
            ("Latitude", lambda s: _safe(s, "lat")),
            ("Longitude", lambda s: _safe(s, "lon")),
            ("Area (ha)", lambda s: _safe(s, "area_ha")),
            ("Proposed Use", lambda s: _safe(s, "proposed_use")),
            ("Capacity (MW)", lambda s: _safe(s, "capacity_mw")),
            ("DNO Area", lambda s: _safe(s, "dno_area")),
        ]),
        ("Grid Connection", [
            ("Headroom (MW)", lambda s: _safe(s, "grid", "headroom_mw")),
            ("Connection Voltage (kV)", lambda s: _safe(s, "grid", "voltage_kv")),
            ("Queue Position", lambda s: _safe(s, "grid", "queue_position")),
            ("Timeline (months)", lambda s: _safe(s, "grid", "timeline_months")),
            ("Cost P10", lambda s: _safe(s, "grid", "cost_p10")),
            ("Cost P50", lambda s: _safe(s, "grid", "cost_p50")),
            ("Cost P90", lambda s: _safe(s, "grid", "cost_p90")),
        ]),
        ("Financial", [
            ("CapEx", lambda s: _safe(s, "financial", "capex")),
            ("OpEx (annual)", lambda s: _safe(s, "financial", "opex_annual")),
            ("Revenue (annual)", lambda s: _safe(s, "financial", "revenue_annual")),
            ("IRR (%)", lambda s: _safe(s, "financial", "irr")),
            ("Payback (years)", lambda s: _safe(s, "financial", "payback_years")),
            ("NPV", lambda s: _safe(s, "financial", "npv")),
            ("LCOE (£/MWh)", lambda s: _safe(s, "financial", "lcoe")),
        ]),
        ("Planning", [
            ("LPA", lambda s: _safe(s, "planning", "lpa")),
            ("Approval Rate (%)", lambda s: _safe(s, "planning", "approval_rate")),
            ("EIA Required", lambda s: _safe(s, "planning", "eia_required")),
            ("BNG Obligation", lambda s: _safe(s, "planning", "bng_obligation")),
            ("Planning Risk", lambda s: _safe(s, "planning", "planning_risk")),
        ]),
        ("Environmental", [
            ("Constraint Count", lambda s: len((_safe(s, "environmental", "constraints") or []) if isinstance(_safe(s, "environmental", "constraints"), list) else [])),
            ("Survey Cost Est.", lambda s: _safe(s, "environmental", "estimated_cost")),
            ("Flood Zone", lambda s: _safe(s, "environmental", "flood_zone")),
        ]),
    ]

    for section_name, metrics in comparison_metrics:
        # Section header
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        cell = ws1.cell(row=row, column=1, value=section_name)
        cell.font = section_font
        cell.fill = section_fill
        for c in range(1, total_cols + 1):
            ws1.cell(row=row, column=c).fill = section_fill
        row += 1

        for metric_name, getter in metrics:
            ws1.cell(row=row, column=1, value=metric_name).font = label_font
            ws1.cell(row=row, column=1).border = thin_border
            for col_idx, site in enumerate(sites, 2):
                val = getter(site)
                cell = ws1.cell(row=row, column=col_idx, value=val)
                cell.font = value_font
                cell.border = thin_border
                if row % 2 == 0:
                    cell.fill = bg_fill
            row += 1

    # Conditional formatting on verdict cells (row 4 = first data row after header)
    verdict_col_start = 2
    for col_idx in range(verdict_col_start, verdict_col_start + num_sites):
        col_l = get_column_letter(col_idx)
        # Search for verdict cells and apply formatting
        for r in range(1, row):
            cell = ws1.cell(row=r, column=1)
            if cell.value == "Verdict":
                for ci in range(2, 2 + num_sites):
                    vc = ws1.cell(row=r, column=ci)
                    v = str(vc.value or "").upper()
                    if v in ("GO", "PROCEED"):
                        vc.fill = green_fill
                    elif v in ("CAUTION", "CONDITIONAL"):
                        vc.fill = amber_fill
                    elif v in ("NO-GO", "NOGO", "NO GO", "REJECT"):
                        vc.fill = red_fill
            if cell.value == "Overall Score":
                for ci in range(2, 2 + num_sites):
                    sc = ws1.cell(row=r, column=ci)
                    try:
                        sv = float(sc.value)
                        if sv >= 70:
                            sc.fill = green_fill
                        elif sv >= 40:
                            sc.fill = amber_fill
                        else:
                            sc.fill = red_fill
                    except (ValueError, TypeError):
                        pass

    _auto_width(ws1)

    # ═══════════════════════════════════════════════════════
    # Sheet 2: Grid Data
    # ═══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Grid Data")
    row = _brand_header(ws2, 1, 8)

    for si, site in enumerate(sites):
        sname = site.get("name", f"Site {si + 1}")
        g = site.get("grid") or {}

        # Site label
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws2.cell(row=row, column=1, value=sname)
        cell.font = section_font
        cell.fill = section_fill
        for c in range(1, 9):
            ws2.cell(row=row, column=c).fill = section_fill
        row += 1

        # Substation table
        subs = g.get("nearest_substations") or []
        if subs:
            _header_row(ws2, row, [
                "Substation", "Voltage (kV)", "Distance (km)", "Headroom (MW)",
                "Queue", "Demand (MW)", "Generation (MW)", "Firm Capacity (MW)"
            ])
            row += 1
            for sub in subs[:8]:
                if isinstance(sub, dict):
                    vals = [
                        sub.get("name", "—"), sub.get("voltage_kv", "—"),
                        sub.get("distance_km", "—"), sub.get("headroom_mw", "—"),
                        sub.get("queue", "—"), sub.get("demand_mw", "—"),
                        sub.get("generation_mw", "—"), sub.get("firm_capacity_mw", "—"),
                    ]
                    for col, v in enumerate(vals, 1):
                        cell = ws2.cell(row=row, column=col, value=v)
                        cell.font = value_font
                        cell.border = thin_border
                        if row % 2 == 0:
                            cell.fill = bg_fill
                    row += 1

        # Summary KVs
        kv_data = [
            ("Total Headroom (MW)", g.get("headroom_mw", "—")),
            ("Connection Voltage (kV)", g.get("voltage_kv", "—")),
            ("Queue Position", g.get("queue_position", "—")),
            ("Timeline (months)", g.get("timeline_months", "—")),
            ("Cost P10", g.get("cost_p10", "—")),
            ("Cost P50", g.get("cost_p50", "—")),
            ("Cost P90", g.get("cost_p90", "—")),
        ]
        for label, val in kv_data:
            ws2.cell(row=row, column=1, value=label).font = label_font
            ws2.cell(row=row, column=1).border = thin_border
            ws2.cell(row=row, column=2, value=val).font = value_font
            ws2.cell(row=row, column=2).border = thin_border
            row += 1

        row += 1  # blank row between sites

    _auto_width(ws2)

    # ═══════════════════════════════════════════════════════
    # Sheet 3: Environmental
    # ═══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Environmental")
    row = _brand_header(ws3, 1, 6)

    for si, site in enumerate(sites):
        sname = site.get("name", f"Site {si + 1}")
        env = site.get("environmental") or {}

        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws3.cell(row=row, column=1, value=sname)
        cell.font = section_font
        cell.fill = section_fill
        for c in range(1, 7):
            ws3.cell(row=row, column=c).fill = section_fill
        row += 1

        constraints = env.get("constraints") or []
        if constraints:
            _header_row(ws3, row, [
                "Constraint Type", "Name", "Distance (m)", "Status", "Detail", "Buffer (m)"
            ])
            row += 1
            for con in constraints[:15]:
                if isinstance(con, dict):
                    vals = [
                        con.get("type", "—"), con.get("name", "—"),
                        con.get("distance_m", "—"),
                        "Blocking" if con.get("blocking") else "Advisory",
                        con.get("detail", "—"), con.get("buffer_m", "—"),
                    ]
                elif isinstance(con, str):
                    vals = [con, "—", "—", "—", "—", "—"]
                else:
                    continue
                for col, v in enumerate(vals, 1):
                    cell = ws3.cell(row=row, column=col, value=v)
                    cell.font = value_font
                    cell.border = thin_border
                    if row % 2 == 0:
                        cell.fill = bg_fill
                    # Colour blocking/advisory
                    if col == 4:
                        cell.fill = red_fill if v == "Blocking" else amber_fill
                row += 1
        else:
            ws3.cell(row=row, column=1,
                     value="No environmental constraints identified").font = value_font
            row += 1

        # Surveys
        surveys = env.get("required_surveys") or []
        if surveys:
            row += 1
            ws3.cell(row=row, column=1, value="Required Surveys").font = label_font
            row += 1
            _header_row(ws3, row, ["Survey", "Est. Cost", "Timeline", "", "", ""])
            row += 1
            for sv in surveys:
                if isinstance(sv, dict):
                    vals = [sv.get("survey", "—"), sv.get("cost", "—"), sv.get("timeline", "—")]
                elif isinstance(sv, str):
                    vals = [sv, "—", "—"]
                else:
                    continue
                for col, v in enumerate(vals, 1):
                    ws3.cell(row=row, column=col, value=v).font = value_font
                    ws3.cell(row=row, column=col).border = thin_border
                row += 1

        row += 1

    _auto_width(ws3)

    # ═══════════════════════════════════════════════════════
    # Sheet 4: Financial
    # ═══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Financial")
    row = _brand_header(ws4, 1, num_sites + 1)

    _header_row(ws4, row, ["Financial Metric"] + site_names)
    row += 1

    fin_metrics = [
        ("CapEx (Total)", lambda s: _safe(s, "financial", "capex")),
        ("OpEx (Annual)", lambda s: _safe(s, "financial", "opex_annual")),
        ("Revenue (Annual)", lambda s: _safe(s, "financial", "revenue_annual")),
        ("IRR (%)", lambda s: _safe(s, "financial", "irr")),
        ("Payback (years)", lambda s: _safe(s, "financial", "payback_years")),
        ("NPV", lambda s: _safe(s, "financial", "npv")),
        ("LCOE (£/MWh)", lambda s: _safe(s, "financial", "lcoe")),
        ("Grid Connection Cost (P50)", lambda s: _safe(s, "grid", "cost_p50")),
        ("Capacity (MW)", lambda s: _safe(s, "capacity_mw")),
    ]

    for metric_name, getter in fin_metrics:
        ws4.cell(row=row, column=1, value=metric_name).font = label_font
        ws4.cell(row=row, column=1).border = thin_border
        for col_idx, site in enumerate(sites, 2):
            val = getter(site)
            cell = ws4.cell(row=row, column=col_idx, value=val)
            cell.font = value_font
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = bg_fill
        row += 1

    # IRR conditional formatting
    row += 1
    ws4.cell(row=row, column=1, value="Colour Key:").font = label_font
    row += 1
    for label, fill in [("IRR >= 8% (Strong)", green_fill),
                         ("IRR 5-8% (Marginal)", amber_fill),
                         ("IRR < 5% (Weak)", red_fill)]:
        ws4.cell(row=row, column=1, value=label).font = value_font
        ws4.cell(row=row, column=1).fill = fill
        row += 1

    # Apply IRR conditional colours
    for r in range(1, row):
        if ws4.cell(row=r, column=1).value == "IRR (%)":
            for ci in range(2, 2 + num_sites):
                cell = ws4.cell(row=r, column=ci)
                try:
                    irr_val = float(cell.value)
                    if irr_val >= 8:
                        cell.fill = green_fill
                    elif irr_val >= 5:
                        cell.fill = amber_fill
                    else:
                        cell.fill = red_fill
                except (ValueError, TypeError):
                    pass

    # CapEx breakdown sub-table per site
    row += 1
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_sites + 1)
    ws4.cell(row=row, column=1, value="CapEx Breakdown by Site").font = section_font
    ws4.cell(row=row, column=1).fill = section_fill
    row += 1

    # Collect all breakdown keys
    all_keys = set()
    for site in sites:
        bd = (site.get("financial") or {}).get("capex_breakdown") or {}
        all_keys.update(bd.keys())
    all_keys = sorted(all_keys)

    if all_keys:
        _header_row(ws4, row, ["Component"] + site_names)
        row += 1
        for key in all_keys:
            ws4.cell(row=row, column=1,
                     value=key.replace("_", " ").title()).font = label_font
            ws4.cell(row=row, column=1).border = thin_border
            for ci, site in enumerate(sites, 2):
                bd = (site.get("financial") or {}).get("capex_breakdown") or {}
                cell = ws4.cell(row=row, column=ci, value=bd.get(key, "—"))
                cell.font = value_font
                cell.border = thin_border
            row += 1

    _auto_width(ws4)

    # ── Save ───────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 3. CSV Export
# ---------------------------------------------------------------------------

def generate_csv_export(sites: list[dict]) -> str:
    """
    Generate a flat CSV string for quick data export.

    One row per site, all key metrics as columns.
    """
    if not sites:
        return ""

    # Define columns: (header, getter)
    columns = [
        ("Site Name", lambda s: _safe(s, "name")),
        ("Verdict", lambda s: _safe(s, "verdict")),
        ("Score", lambda s: _safe(s, "score")),
        ("Latitude", lambda s: _safe(s, "lat")),
        ("Longitude", lambda s: _safe(s, "lon")),
        ("Area (ha)", lambda s: _safe(s, "area_ha")),
        ("Proposed Use", lambda s: _safe(s, "proposed_use")),
        ("Capacity (MW)", lambda s: _safe(s, "capacity_mw")),
        ("DNO Area", lambda s: _safe(s, "dno_area")),
        ("Grid Headroom (MW)", lambda s: _safe(s, "grid", "headroom_mw")),
        ("Grid Voltage (kV)", lambda s: _safe(s, "grid", "voltage_kv")),
        ("Queue Position", lambda s: _safe(s, "grid", "queue_position")),
        ("Grid Timeline (months)", lambda s: _safe(s, "grid", "timeline_months")),
        ("Connection Cost P10", lambda s: _safe(s, "grid", "cost_p10")),
        ("Connection Cost P50", lambda s: _safe(s, "grid", "cost_p50")),
        ("Connection Cost P90", lambda s: _safe(s, "grid", "cost_p90")),
        ("Constraint Count", lambda s: len(
            _safe(s, "environmental", "constraints")
            if isinstance(_safe(s, "environmental", "constraints"), list) else []
        )),
        ("Env Survey Cost", lambda s: _safe(s, "environmental", "estimated_cost")),
        ("Flood Zone", lambda s: _safe(s, "environmental", "flood_zone")),
        ("CapEx", lambda s: _safe(s, "financial", "capex")),
        ("OpEx (annual)", lambda s: _safe(s, "financial", "opex_annual")),
        ("Revenue (annual)", lambda s: _safe(s, "financial", "revenue_annual")),
        ("IRR (%)", lambda s: _safe(s, "financial", "irr")),
        ("Payback (years)", lambda s: _safe(s, "financial", "payback_years")),
        ("NPV", lambda s: _safe(s, "financial", "npv")),
        ("LCOE (£/MWh)", lambda s: _safe(s, "financial", "lcoe")),
        ("LPA", lambda s: _safe(s, "planning", "lpa")),
        ("Approval Rate (%)", lambda s: _safe(s, "planning", "approval_rate")),
        ("EIA Required", lambda s: _safe(s, "planning", "eia_required")),
        ("BNG Obligation", lambda s: _safe(s, "planning", "bng_obligation")),
        ("Planning Risk", lambda s: _safe(s, "planning", "planning_risk")),
        ("Recommendation", lambda s: _safe(s, "recommendation")),
    ]

    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow([c[0] for c in columns])

    # Data rows
    for site in sites:
        writer.writerow([c[1](site) for c in columns])

    return output.getvalue()
